"""
five_prong_report_engine.py -- VSEPR-SIM Back-Matter & Figure Generation
========================================================================

Reads the 5-prong CSV data produced by five_prong_data_gen.py and generates:

  Phase A: High-DPI scientific figures (PNG 300 DPI)
      - Per-material thermophysical grids (T-h, T-s, T-Cp, P-v, T-Z diagrams)
      - Metal property comparison charts (bar, radar, scatter)
      - Heating Cp-vs-T curves per metal and overlay
      - Alloy property comparison charts
      - Crystal catalog distribution charts
      - SmartSampling score distributions and tier comparisons
      - Golden Project cross-pillar rankings

  Phase B: Molecular structure renderings
      - 3D ball-and-stick renders from XYZ library
      - 2D topology projections
      - Bond-diagram schematics

  Phase C: Consolidated report assembly
      - Multi-page XLSX with embedded image sheets
      - LaTeX back-matter document (figures + data tables)
      - Markdown figure index with provenance

Every figure is deterministic: same data → identical PNG output.
Anti-black-box: every metric plotted traces to a CSV row.

Usage:
    python five_prong_report_engine.py
    python five_prong_report_engine.py --run-dir out/five_prong_run
    python five_prong_report_engine.py --skip-molecules
    python five_prong_report_engine.py --skip-latex

VSEPR-SIM 4.0.4.03
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import time
import glob
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import Normalize
from matplotlib.backends.backend_pdf import PdfPages

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SCRIPT_DIR
sys.path.insert(0, str(_PROJECT_ROOT))

from pykernel.chart_helpers import (
    configure_style, save_figure, PALETTE, PALETTE_CYCLE,
    chart_line, chart_scatter, chart_bar, chart_barh,
    chart_histogram, chart_heatmap, chart_radar, chart_box,
    chart_stacked_area, chart_pie,
    render_molecule_3d, render_molecule_2d, render_bond_diagram,
    parse_xyz, parse_multi_xyz,
    latex_figure, latex_subfigure_pair, latex_subfigure_grid, latex_table,
)
from pykernel.species_shell import (
    run_batch_export as species_batch_export,
    run_batch_figures as species_batch_figures,
    SpeciesShell,
)

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════

DPI = 300
FIG_W, FIG_H = 10, 7
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
SECTION_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)


def _log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    print(f"[{ts}] {msg}", flush=True)


# ═══════════════════════════════════════════════════════════════════════
# CSV LOADING
# ═══════════════════════════════════════════════════════════════════════

def _load_csv(path: Path) -> List[dict]:
    if not path.exists():
        return []
    with open(path, "r", newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _safe_float(val, default=0.0):
    if val is None or val == "" or val == "None":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════════════════════════════════
# PHASE A: SCIENTIFIC FIGURE GENERATION
# ═══════════════════════════════════════════════════════════════════════

class FigureEngine:
    """Generates all scientific figures from 5-prong CSV data."""

    def __init__(self, run_dir: Path, fig_dir: Path):
        self.run_dir = run_dir
        self.fig_dir = fig_dir
        self.fig_dir.mkdir(parents=True, exist_ok=True)
        self.manifest: List[dict] = []
        configure_style()

    def _record(self, name: str, category: str, title: str,
                description: str = "", data_source: str = ""):
        self.manifest.append({
            "filename": name,
            "category": category,
            "title": title,
            "description": description,
            "data_source": data_source,
        })

    # ───────────────────────────────────────────────────────────────
    # A1: Thermophysical Atlas Figures
    # ───────────────────────────────────────────────────────────────

    def generate_thermo_figures(self) -> int:
        _log("Phase A1: Thermophysical Atlas figures")
        grid_data = _load_csv(self.run_dir / "prong1_state_grid.csv")
        if not grid_data:
            _log("  No thermophysical grid data found, skipping")
            return 0

        by_material = defaultdict(list)
        for row in grid_data:
            by_material[row["material"]].append(row)

        count = 0
        materials = sorted(by_material.keys())

        # Per-material property diagrams
        for mat in materials:
            rows = by_material[mat]
            T = np.array([_safe_float(r["T_K"]) for r in rows])
            P = np.array([_safe_float(r["P_Pa"]) for r in rows])
            h = np.array([_safe_float(r["h_Jkg"]) for r in rows])
            s = np.array([_safe_float(r["s_JkgK"]) for r in rows])
            cp = np.array([_safe_float(r["cp_JkgK"]) for r in rows])
            Z = np.array([_safe_float(r["Z_comp"]) for r in rows])

            valid = (T > 0) & (h != 0)
            if valid.sum() < 10:
                continue

            # T-h diagram
            fig = chart_scatter(
                T[valid], h[valid], colors=P[valid],
                title=f"{mat} — Temperature vs Enthalpy",
                xlabel="Temperature (K)", ylabel="Enthalpy (J/kg)",
                colorbar_label="Pressure (Pa)", cmap="plasma",
            )
            name = f"thermo_{mat}_T_h.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "thermo", f"{mat}: T-h Diagram",
                         "Temperature vs enthalpy coloured by pressure",
                         "prong1_state_grid.csv")
            count += 1

            # T-s diagram
            sv = (T > 0) & (s != 0)
            if sv.sum() > 10:
                fig = chart_scatter(
                    T[sv], s[sv], colors=P[sv],
                    title=f"{mat} — Temperature vs Entropy",
                    xlabel="Temperature (K)", ylabel="Entropy (J/kg·K)",
                    colorbar_label="Pressure (Pa)", cmap="viridis",
                )
                name = f"thermo_{mat}_T_s.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "thermo", f"{mat}: T-s Diagram",
                             data_source="prong1_state_grid.csv")
                count += 1

            # T-Cp diagram
            cpv = (T > 0) & (cp > 0)
            if cpv.sum() > 10:
                fig = chart_scatter(
                    T[cpv], cp[cpv], colors=P[cpv],
                    title=f"{mat} — Isobaric Heat Capacity",
                    xlabel="Temperature (K)", ylabel="Cp (J/kg·K)",
                    colorbar_label="Pressure (Pa)", cmap="inferno",
                )
                name = f"thermo_{mat}_T_Cp.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "thermo", f"{mat}: T-Cp Diagram",
                             data_source="prong1_state_grid.csv")
                count += 1

            # Compressibility factor heatmap
            if Z.max() > 0:
                unique_T = np.unique(T)
                unique_P = np.unique(P)
                if len(unique_T) > 2 and len(unique_P) > 2:
                    n_T = min(len(unique_T), 80)
                    n_P = min(len(unique_P), 80)
                    T_edges = np.linspace(T.min(), T.max(), n_T)
                    P_edges = np.linspace(P.min(), P.max(), n_P)
                    Z_grid = np.full((n_P - 1, n_T - 1), np.nan)
                    for r in rows:
                        t_val = _safe_float(r["T_K"])
                        p_val = _safe_float(r["P_Pa"])
                        z_val = _safe_float(r["Z_comp"])
                        ti = np.searchsorted(T_edges, t_val) - 1
                        pi = np.searchsorted(P_edges, p_val) - 1
                        ti = max(0, min(ti, n_T - 2))
                        pi = max(0, min(pi, n_P - 2))
                        Z_grid[pi, ti] = z_val

                    fig = chart_heatmap(
                        Z_grid,
                        title=f"{mat} — Compressibility Factor Z(T, P)",
                        colorbar_label="Z", cmap="coolwarm",
                    )
                    name = f"thermo_{mat}_Z_heatmap.png"
                    save_figure(fig, name, outdir=self.fig_dir)
                    self._record(name, "thermo", f"{mat}: Z Heatmap",
                                 data_source="prong1_state_grid.csv")
                    count += 1

        # Cross-material overlay: T-Cp for first 8 materials
        overlay_mats = materials[:8]
        if len(overlay_mats) > 1:
            x_arrs, y_arrs, labels = [], [], []
            for mat in overlay_mats:
                rows = by_material[mat]
                T = np.array([_safe_float(r["T_K"]) for r in rows])
                cp = np.array([_safe_float(r["cp_JkgK"]) for r in rows])
                mask = (T > 0) & (cp > 0)
                if mask.sum() > 5:
                    idx = np.argsort(T[mask])
                    # Take every nth point for readability
                    step = max(1, len(idx) // 200)
                    x_arrs.append(T[mask][idx][::step])
                    y_arrs.append(cp[mask][idx][::step])
                    labels.append(mat)
            if len(x_arrs) > 1:
                fig = chart_line(
                    x_arrs, y_arrs, labels=labels,
                    title="Cross-Material Cp Comparison",
                    xlabel="Temperature (K)", ylabel="Cp (J/kg·K)",
                )
                name = "thermo_cross_Cp_overlay.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "thermo", "Cross-Material Cp Overlay",
                             data_source="prong1_state_grid.csv")
                count += 1

        _log(f"  Thermophysical figures: {count}")
        return count

    # ───────────────────────────────────────────────────────────────
    # A2: Crystal Discovery Figures
    # ───────────────────────────────────────────────────────────────

    def generate_crystal_figures(self) -> int:
        _log("Phase A2: Crystal Discovery figures")
        count = 0

        catalog = _load_csv(self.run_dir / "prong2_catalog.csv")
        subs = _load_csv(self.run_dir / "prong2_substitutions.csv")
        perts = _load_csv(self.run_dir / "prong2_perturbations.csv")

        # Crystal class distribution pie
        if catalog:
            class_counts = defaultdict(int)
            for row in catalog:
                class_counts[row.get("crystal_class", "unknown")] += 1
            fig = chart_pie(
                list(class_counts.values()), list(class_counts.keys()),
                title="Crystal Catalog: Class Distribution",
            )
            name = "crystal_class_distribution.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "crystal", "Crystal Class Distribution",
                         data_source="prong2_catalog.csv")
            count += 1

        # Bravais lattice distribution bar
        if catalog:
            bravais_counts = defaultdict(int)
            for row in catalog:
                bravais_counts[row.get("bravais", "unknown")] += 1
            fig = chart_bar(
                list(bravais_counts.keys()),
                list(bravais_counts.values()),
                title="Crystal Catalog: Bravais Lattice Types",
                ylabel="Count",
            )
            name = "crystal_bravais_distribution.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "crystal", "Bravais Lattice Distribution",
                         data_source="prong2_catalog.csv")
            count += 1

        # Density vs lattice parameter scatter
        if catalog:
            a_vals = np.array([_safe_float(r.get("a_A")) for r in catalog])
            d_vals = np.array([_safe_float(r.get("density_gcc")) for r in catalog])
            valid = (a_vals > 0) & (d_vals > 0)
            if valid.sum() > 3:
                fig = chart_scatter(
                    a_vals[valid], d_vals[valid],
                    title="Crystal Lattice Parameter vs Density",
                    xlabel="Lattice Parameter a (Å)",
                    ylabel="Density (g/cm³)",
                )
                name = "crystal_a_vs_density.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "crystal", "Lattice Parameter vs Density",
                             data_source="prong2_catalog.csv")
                count += 1

        # Stability score distribution for substitutions
        if subs:
            scores = np.array([_safe_float(r.get("stability_score")) for r in subs])
            if len(scores) > 5:
                fig = chart_histogram(
                    scores, bins=40,
                    title="Substitution Scan: Stability Score Distribution",
                    xlabel="Stability Score", ylabel="Count",
                )
                name = "crystal_sub_stability_hist.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "crystal", "Substitution Stability Distribution",
                             data_source="prong2_substitutions.csv")
                count += 1

        # Stability class breakdown bar for substitutions
        if subs:
            class_counts = defaultdict(int)
            for r in subs:
                class_counts[r.get("stability_class", "unknown")] += 1
            fig = chart_bar(
                list(class_counts.keys()),
                list(class_counts.values()),
                title="Substitution Stability Class Breakdown",
                ylabel="Count",
            )
            name = "crystal_sub_class_bar.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "crystal", "Substitution Stability Classes",
                         data_source="prong2_substitutions.csv")
            count += 1

        # Perturbation stability histogram
        if perts:
            scores = np.array([_safe_float(r.get("stability_score")) for r in perts])
            if len(scores) > 5:
                fig = chart_histogram(
                    scores, bins=30,
                    title="Stoichiometric Perturbation: Stability Score Distribution",
                    xlabel="Stability Score", ylabel="Count",
                )
                name = "crystal_pert_stability_hist.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "crystal",
                             "Perturbation Stability Distribution",
                             data_source="prong2_perturbations.csv")
                count += 1

        _log(f"  Crystal discovery figures: {count}")
        return count

    # ───────────────────────────────────────────────────────────────
    # A3: Metals & Macros Figures
    # ───────────────────────────────────────────────────────────────

    def generate_metals_figures(self) -> int:
        _log("Phase A3: Metals & Macros figures")
        count = 0

        metals = _load_csv(self.run_dir / "prong3_metals.csv")
        alloys = _load_csv(self.run_dir / "prong3_alloys.csv")
        heating = _load_csv(self.run_dir / "prong3_heating.csv")

        if not metals:
            _log("  No metals data found, skipping")
            return 0

        syms = [r["symbol"] for r in metals]
        names = [r["name"] for r in metals]

        # Young's modulus ranking
        E_vals = [_safe_float(r.get("E_GPa")) for r in metals]
        fig = chart_barh(
            syms, E_vals,
            title="Young's Modulus by Element",
            xlabel="E (GPa)",
        )
        name = "metals_E_ranking.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Young's Modulus Ranking",
                     data_source="prong3_metals.csv")
        count += 1

        # Density ranking
        rho_vals = [_safe_float(r.get("density_gcc")) for r in metals]
        fig = chart_barh(
            syms, rho_vals,
            title="Density by Element",
            xlabel="Density (g/cm³)",
        )
        name = "metals_density_ranking.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Density Ranking",
                     data_source="prong3_metals.csv")
        count += 1

        # Thermal conductivity ranking
        k_vals = [_safe_float(r.get("k_WmK")) for r in metals]
        fig = chart_barh(
            syms, k_vals,
            title="Thermal Conductivity by Element",
            xlabel="k (W/m·K)",
        )
        name = "metals_k_ranking.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Thermal Conductivity Ranking",
                     data_source="prong3_metals.csv")
        count += 1

        # Melting point ranking
        Tm_vals = [_safe_float(r.get("Tm_K")) for r in metals]
        fig = chart_barh(
            syms, Tm_vals,
            title="Melting Point by Element",
            xlabel="Tm (K)",
        )
        name = "metals_Tm_ranking.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Melting Point Ranking",
                     data_source="prong3_metals.csv")
        count += 1

        # Yield strength ranking
        Sy_vals = [_safe_float(r.get("Sy_MPa")) for r in metals]
        fig = chart_barh(
            syms, Sy_vals,
            title="Yield Strength by Element",
            xlabel="σ_y (MPa)",
        )
        name = "metals_Sy_ranking.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Yield Strength Ranking",
                     data_source="prong3_metals.csv")
        count += 1

        # E vs density scatter (Ashby-style)
        E_arr = np.array(E_vals)
        rho_arr = np.array(rho_vals)
        valid = (E_arr > 0) & (rho_arr > 0)
        if valid.sum() > 3:
            fig = chart_scatter(
                rho_arr[valid], E_arr[valid],
                labels=[syms[i] for i in range(len(syms)) if valid[i]],
                title="Ashby Plot: Young's Modulus vs Density",
                xlabel="Density (g/cm³)", ylabel="E (GPa)",
            )
            name = "metals_ashby_E_rho.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "metals", "Ashby E-ρ Plot",
                         data_source="prong3_metals.csv")
            count += 1

        # Strength vs melting point scatter
        Sy_arr = np.array(Sy_vals)
        Tm_arr = np.array(Tm_vals)
        valid = (Sy_arr > 0) & (Tm_arr > 0)
        if valid.sum() > 3:
            fig = chart_scatter(
                Tm_arr[valid], Sy_arr[valid],
                labels=[syms[i] for i in range(len(syms)) if valid[i]],
                title="Yield Strength vs Melting Point",
                xlabel="Melting Point (K)", ylabel="σ_y (MPa)",
            )
            name = "metals_Sy_vs_Tm.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "metals", "Strength vs Melting Point",
                         data_source="prong3_metals.csv")
            count += 1

        # Radar charts for top metals (Fe, Al, Ti, Cu, Ni, W)
        radar_metals = ["Fe", "Al", "Ti", "Cu", "Ni", "W"]
        props_for_radar = ["E_GPa", "Sy_MPa", "k_WmK", "Tm_K", "density_gcc"]
        prop_labels = ["Modulus", "Strength", "Conductivity", "Melting Pt", "Density"]

        # Normalize each property to 0–1 across all metals
        prop_arrays = {}
        for prop in props_for_radar:
            arr = np.array([_safe_float(r.get(prop)) for r in metals])
            mx = arr.max() if arr.max() > 0 else 1.0
            prop_arrays[prop] = arr / mx

        metal_lookup = {r["symbol"]: i for i, r in enumerate(metals)}
        for sym in radar_metals:
            if sym not in metal_lookup:
                continue
            idx = metal_lookup[sym]
            values = [float(prop_arrays[p][idx]) for p in props_for_radar]
            fig = chart_radar(
                prop_labels, values,
                title=f"{sym} — Material Property Radar",
                color=PALETTE_CYCLE[radar_metals.index(sym) % len(PALETTE_CYCLE)],
            )
            name = f"metals_radar_{sym}.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "metals", f"{sym} Property Radar",
                         data_source="prong3_metals.csv")
            count += 1

        # Multi-property box plots
        box_props = ["E_GPa", "Sy_MPa", "k_WmK", "density_gcc"]
        box_labels = ["E (GPa)", "σy (MPa)", "k (W/m·K)", "ρ (g/cm³)"]
        box_data = []
        for prop in box_props:
            arr = np.array([_safe_float(r.get(prop)) for r in metals])
            arr = arr[arr > 0]
            if len(arr) > 0:
                box_data.append(arr)
            else:
                box_data.append(np.array([0]))
        fig = chart_box(
            box_data, group_names=box_labels,
            title="Metal Property Distributions",
            ylabel="Value",
        )
        name = "metals_property_boxplots.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "metals", "Property Distribution Box Plots",
                     data_source="prong3_metals.csv")
        count += 1

        # Corrosion class breakdown pie
        corr_counts = defaultdict(int)
        for r in metals:
            corr_counts[r.get("corrosion_class", "unknown")] += 1
        if corr_counts:
            fig = chart_pie(
                list(corr_counts.values()), list(corr_counts.keys()),
                title="Corrosion Class Distribution",
            )
            name = "metals_corrosion_pie.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "metals", "Corrosion Class Distribution",
                         data_source="prong3_metals.csv")
            count += 1

        # ── Heating curves ──
        if heating:
            by_mat = defaultdict(list)
            for row in heating:
                by_mat[row["material"]].append(row)

            # Per-metal Cp vs T curves
            for mat in sorted(by_mat.keys()):
                rows = by_mat[mat]
                T = np.array([_safe_float(r["T_K"]) for r in rows])
                Cp = np.array([_safe_float(r["Cp_at_T"]) for r in rows])
                valid = (T > 0) & (Cp > 0)
                if valid.sum() < 3:
                    continue
                idx = np.argsort(T[valid])
                fig = chart_line(
                    [T[valid][idx]], [Cp[valid][idx]],
                    labels=[mat],
                    title=f"{mat} — Heat Capacity vs Temperature",
                    xlabel="Temperature (K)", ylabel="Cp (J/mol·K)",
                )
                name = f"heating_{mat}_Cp.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "heating", f"{mat}: Cp(T) Curve",
                             data_source="prong3_heating.csv")
                count += 1

            # Overlay of all heating curves
            x_arrs, y_arrs, labels = [], [], []
            for mat in sorted(by_mat.keys()):
                rows = by_mat[mat]
                T = np.array([_safe_float(r["T_K"]) for r in rows])
                Cp = np.array([_safe_float(r["Cp_at_T"]) for r in rows])
                valid = (T > 0) & (Cp > 0)

                if valid.sum() < 3:
                    continue
                idx = np.argsort(T[valid])
                x_arrs.append(T[valid][idx])
                y_arrs.append(Cp[valid][idx])
                labels.append(mat)
            if len(x_arrs) > 1:
                n_series = len(x_arrs)
                cycle_colors = [PALETTE_CYCLE[i % len(PALETTE_CYCLE)]
                                for i in range(n_series)]
                fig = chart_line(
                    x_arrs, y_arrs, labels=labels,
                    colors=cycle_colors,
                    title="All Metals — Cp vs Temperature Overlay",
                    xlabel="Temperature (K)", ylabel="Cp (J/mol·K)",
                    figsize=(14, 9),
                )
                name = "heating_all_overlay.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "heating", "All Metals Cp Overlay",
                             data_source="prong3_heating.csv")
                count += 1

        # ── Alloy property charts ──
        if alloys:
            alloy_names = [r.get("name", r.get("formula", "?")) for r in alloys]
            alloy_E = [_safe_float(r.get("elastic_modulus")) for r in alloys]
            alloy_rho = [_safe_float(r.get("density")) for r in alloys]
            alloy_k = [_safe_float(r.get("thermal_conductivity")) for r in alloys]

            # Alloy E comparison
            fig = chart_barh(
                alloy_names, alloy_E,
                title="Alloy Elastic Modulus (ROM Estimate)",
                xlabel="E (GPa)",
                figsize=(12, max(8, len(alloy_names) * 0.35)),
            )
            name = "alloy_E_comparison.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "alloy", "Alloy Elastic Modulus",
                         data_source="prong3_alloys.csv")
            count += 1

            # Alloy E vs density scatter
            E_arr = np.array(alloy_E)
            rho_arr = np.array(alloy_rho)
            valid = (E_arr > 0) & (rho_arr > 0)
            if valid.sum() > 2:
                fig = chart_scatter(
                    rho_arr[valid], E_arr[valid],
                    labels=[alloy_names[i] for i in range(len(alloy_names)) if valid[i]],
                    title="Alloy Ashby: E vs Density",
                    xlabel="Density (g/cm³)", ylabel="E (GPa)",
                    figsize=(12, 9),
                )
                name = "alloy_ashby_E_rho.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "alloy", "Alloy Ashby E-ρ Plot",
                             data_source="prong3_alloys.csv")
                count += 1

            # Alloy thermal conductivity bar
            fig = chart_barh(
                alloy_names, alloy_k,
                title="Alloy Thermal Conductivity (ROM Estimate)",
                xlabel="k (W/m·K)",
                figsize=(12, max(8, len(alloy_names) * 0.35)),
            )
            name = "alloy_k_comparison.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "alloy", "Alloy Thermal Conductivity",
                         data_source="prong3_alloys.csv")
            count += 1

        _log(f"  Metals & Macros figures: {count}")
        return count

    # ───────────────────────────────────────────────────────────────
    # A4: SmartSampling Figures
    # ───────────────────────────────────────────────────────────────

    def generate_sampling_figures(self) -> int:
        _log("Phase A4: SmartSampling figures")
        count = 0

        samples = _load_csv(self.run_dir / "prong4_sampling.csv")
        if not samples:
            _log("  No sampling data found, skipping")
            return 0

        # Total score distribution histogram
        scores = np.array([_safe_float(r.get("total_score")) for r in samples])
        fig = chart_histogram(
            scores, bins=50,
            title="SmartSampling: Total Score Distribution",
            xlabel="Total Score", ylabel="Count",
        )
        name = "sampling_score_distribution.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "sampling", "Total Score Distribution",
                     data_source="prong4_sampling.csv")
        count += 1

        # Score components box plot
        components = ["novelty", "objective", "uncertainty",
                      "plausibility", "instability", "cost"]
        comp_data = []
        for c in components:
            arr = np.array([_safe_float(r.get(c)) for r in samples])
            comp_data.append(arr)
        fig = chart_box(
            comp_data, group_names=components,
            title="SmartSampling: Score Component Distributions",
            ylabel="Score",
        )
        name = "sampling_component_boxplot.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "sampling", "Score Component Box Plots",
                     data_source="prong4_sampling.csv")
        count += 1

        # Per-domain score comparison
        by_domain = defaultdict(list)
        for r in samples:
            domain = r.get("domain", "unknown")
            if domain.startswith("steam"):
                by_domain["steam"].append(r)
            elif domain == "crystal":
                by_domain["crystal"].append(r)
            elif domain.startswith("alloy"):
                by_domain["alloy"].append(r)
            else:
                by_domain[domain].append(r)

        if len(by_domain) > 1:
            dom_names = sorted(by_domain.keys())
            dom_means = [np.mean([_safe_float(r.get("total_score"))
                                  for r in by_domain[d]])
                         for d in dom_names]
            fig = chart_bar(
                dom_names, dom_means,
                title="SmartSampling: Mean Score by Domain",
                ylabel="Mean Total Score",
            )
            name = "sampling_domain_comparison.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "sampling", "Mean Score by Domain",
                         data_source="prong4_sampling.csv")
            count += 1

        # Tier distribution pie
        tier_counts = defaultdict(int)
        for r in samples:
            tier_counts[r.get("tier", "unknown")] += 1
        if tier_counts:
            fig = chart_pie(
                list(tier_counts.values()), list(tier_counts.keys()),
                title="SmartSampling: Tier Distribution",
            )
            name = "sampling_tier_pie.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "sampling", "Tier Distribution",
                         data_source="prong4_sampling.csv")
            count += 1

        # Novelty vs objective scatter
        nov = np.array([_safe_float(r.get("novelty")) for r in samples])
        obj = np.array([_safe_float(r.get("objective")) for r in samples])
        tot = np.array([_safe_float(r.get("total_score")) for r in samples])
        fig = chart_scatter(
            nov, obj, colors=tot,
            title="SmartSampling: Novelty vs Objective Match",
            xlabel="Novelty Score", ylabel="Objective Match",
            colorbar_label="Total Score", cmap="plasma",
        )
        name = "sampling_novelty_vs_objective.png"
        save_figure(fig, name, outdir=self.fig_dir)
        self._record(name, "sampling", "Novelty vs Objective",
                     data_source="prong4_sampling.csv")
        count += 1

        # Per-material score distribution for steam materials
        steam_samples = [r for r in samples if r.get("domain", "").startswith("steam")]
        if steam_samples:
            by_mat = defaultdict(list)
            for r in steam_samples:
                by_mat[r["material"]].append(_safe_float(r.get("total_score")))
            mat_names = sorted(by_mat.keys())
            mat_data = [np.array(by_mat[m]) for m in mat_names]
            if len(mat_data) > 1:
                fig = chart_box(
                    mat_data, group_names=mat_names,
                    title="Steam Sampling: Score by Material",
                    ylabel="Total Score",
                )
                name = "sampling_steam_by_material.png"
                save_figure(fig, name, outdir=self.fig_dir)
                self._record(name, "sampling",
                             "Steam Sampling Score by Material",
                             data_source="prong4_sampling.csv")
                count += 1

        _log(f"  SmartSampling figures: {count}")
        return count

    # ───────────────────────────────────────────────────────────────
    # A5: Golden Project Figures
    # ───────────────────────────────────────────────────────────────

    def generate_golden_figures(self) -> int:
        _log("Phase A5: Golden Project figures")
        count = 0

        results_csv = self.run_dir / "golden_project" / "discovery_results.csv"
        results = _load_csv(results_csv)
        if not results:
            _log("  No golden project results found, skipping")
            return 0

        # Results per pillar bar chart
        by_pillar = defaultdict(int)
        for r in results:
            by_pillar[r.get("pillar", "unknown")] += 1
        if by_pillar:
            fig = chart_bar(
                list(by_pillar.keys()),
                list(by_pillar.values()),
                title="Golden Project: Results by Pillar",
                ylabel="Result Count",
            )
            name = "golden_pillar_distribution.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "golden", "Results by Pillar",
                         data_source="golden_project/discovery_results.csv")
            count += 1

        # Score distribution histogram
        scores = np.array([_safe_float(r.get("score")) for r in results])
        if len(scores) > 5:
            fig = chart_histogram(
                scores, bins=50,
                title="Golden Project: Score Distribution",
                xlabel="Score", ylabel="Count",
            )
            name = "golden_score_histogram.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "golden", "Score Distribution",
                         data_source="golden_project/discovery_results.csv")
            count += 1

        # Per-mission result count bar
        by_mission = defaultdict(int)
        for r in results:
            by_mission[r.get("mission_id", "unknown")] += 1
        if by_mission:
            fig = chart_barh(
                list(by_mission.keys()),
                list(by_mission.values()),
                title="Golden Project: Results per Mission",
                xlabel="Count",
                figsize=(12, max(6, len(by_mission) * 0.5)),
            )
            name = "golden_mission_counts.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "golden", "Results per Mission",
                         data_source="golden_project/discovery_results.csv")
            count += 1

        # Top-ranked materials scatter
        top = [r for r in results if _safe_float(r.get("rank")) <= 10]
        if top:
            mats = [r.get("material", "?") for r in top]
            sc = np.array([_safe_float(r.get("score")) for r in top])
            rk = np.array([_safe_float(r.get("rank")) for r in top])
            fig = chart_scatter(
                rk, sc, labels=mats,
                title="Golden Project: Top-10 Results",
                xlabel="Rank", ylabel="Score",
            )
            name = "golden_top10_scatter.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "golden", "Top-10 Ranked Results",
                         data_source="golden_project/discovery_results.csv")
            count += 1

        # Material frequency bar (most frequently appearing materials)
        mat_freq = defaultdict(int)
        for r in results:
            mat = r.get("material", "")
            if mat:
                mat_freq[mat] += 1
        if mat_freq:
            top_mats = sorted(mat_freq.items(), key=lambda x: -x[1])[:20]
            fig = chart_barh(
                [m for m, _ in top_mats],
                [c for _, c in top_mats],
                title="Golden Project: Most Frequent Materials",
                xlabel="Occurrences",
                figsize=(12, max(7, len(top_mats) * 0.4)),
            )
            name = "golden_material_frequency.png"
            save_figure(fig, name, outdir=self.fig_dir)
            self._record(name, "golden", "Material Frequency",
                         data_source="golden_project/discovery_results.csv")
            count += 1

        _log(f"  Golden Project figures: {count}")
        return count


# ═══════════════════════════════════════════════════════════════════════
# PHASE B: MOLECULAR STRUCTURE RENDERINGS
# ═══════════════════════════════════════════════════════════════════════

class MoleculeRenderer:
    """Renders 3D/2D molecular structures from XYZ files."""

    def __init__(self, xyz_dir: Path, fig_dir: Path, max_molecules: int = 50):
        self.xyz_dir = xyz_dir
        self.fig_dir = fig_dir / "molecules"
        self.fig_dir.mkdir(parents=True, exist_ok=True)
        self.max_molecules = max_molecules
        self.manifest: List[dict] = []

    def _record(self, name: str, title: str, description: str = ""):
        self.manifest.append({
            "filename": name,
            "category": "molecule",
            "title": title,
            "description": description,
            "data_source": "XYZ library",
        })

    def render_all(self) -> int:
        _log("Phase B: Molecular structure renderings")
        if not self.xyz_dir.exists():
            _log(f"  XYZ directory not found: {self.xyz_dir}")
            return 0

        xyz_files = sorted(self.xyz_dir.glob("*.xyz"))[:self.max_molecules]
        if not xyz_files:
            _log("  No XYZ files found")
            return 0

        _log(f"  XYZ files found: {len(xyz_files)}")
        count = 0

        for xyz_path in xyz_files:
            try:
                symbols, positions, title = parse_xyz(xyz_path)
                if len(symbols) < 2:
                    continue
                mol_name = xyz_path.stem

                # 3D ball-and-stick
                name_3d = f"mol3d_{mol_name}.png"
                render_molecule_3d(
                    symbols, positions,
                    title=title or mol_name,
                    outpath=self.fig_dir / name_3d,
                    dpi=DPI,
                )
                self._record(name_3d, f"{mol_name} — 3D Structure",
                             f"{len(symbols)} atoms, 3D ball-and-stick")
                count += 1
                plt.close("all")

                # 2D topology
                name_2d = f"mol2d_{mol_name}.png"
                render_molecule_2d(
                    symbols, positions,
                    title=title or mol_name,
                    outpath=self.fig_dir / name_2d,
                    dpi=DPI,
                )
                self._record(name_2d, f"{mol_name} — 2D Topology",
                             f"{len(symbols)} atoms, XY projection")
                count += 1
                plt.close("all")

                # Bond diagram
                name_bd = f"molbd_{mol_name}.png"
                render_bond_diagram(
                    symbols, positions,
                    title=title or mol_name,
                    outpath=self.fig_dir / name_bd,
                    dpi=DPI,
                )
                self._record(name_bd, f"{mol_name} — Bond Diagram",
                             f"{len(symbols)} atoms, topology schematic")
                count += 1
                plt.close("all")

            except Exception as e:
                _log(f"  WARN: Failed to render {xyz_path.name}: {e}")
                plt.close("all")
                continue

        _log(f"  Molecular renderings: {count}")
        return count


# ═══════════════════════════════════════════════════════════════════════
# PHASE C1: PDF MULTI-PAGE FIGURE ATLAS
# ═══════════════════════════════════════════════════════════════════════

def build_pdf_atlas(fig_dir: Path, output_path: Path,
                    manifest: List[dict]) -> str:
    """Compile all generated PNGs into a single multi-page PDF atlas."""
    _log("Phase C1: Building PDF figure atlas")

    categories = defaultdict(list)
    for entry in manifest:
        categories[entry["category"]].append(entry)

    category_order = ["thermo", "crystal", "metals", "heating",
                      "alloy", "sampling", "golden", "molecule"]

    with PdfPages(str(output_path)) as pdf:
        # Title page
        fig, ax = plt.subplots(figsize=(11, 8.5))
        ax.axis("off")
        ax.text(0.5, 0.65, "VSEPR-SIM", fontsize=42, fontweight="bold",
                ha="center", va="center", color="#2e86c1")
        ax.text(0.5, 0.50, "5-Prong Data Generation", fontsize=28,
                ha="center", va="center", color="#2c3e50")
        ax.text(0.5, 0.40, "Back-Matter: Figures & Renderings",
                fontsize=20, ha="center", va="center", color="#7f8c8d")
        ax.text(0.5, 0.28, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                fontsize=14, ha="center", va="center", color="#7f8c8d")
        ax.text(0.5, 0.22, f"Total Figures: {len(manifest)}",
                fontsize=14, ha="center", va="center", color="#7f8c8d")
        pdf.savefig(fig)
        plt.close(fig)

        # Table of contents page
        fig, ax = plt.subplots(figsize=(11, 8.5))
        ax.axis("off")
        ax.text(0.5, 0.95, "Table of Contents", fontsize=24,
                fontweight="bold", ha="center", va="top", color="#2c3e50")
        y = 0.85
        for cat in category_order:
            if cat in categories:
                entries = categories[cat]
                cat_title = cat.replace("_", " ").title()
                ax.text(0.1, y, f"■ {cat_title}  ({len(entries)} figures)",
                        fontsize=13, fontweight="bold", color="#2e86c1",
                        va="top")
                y -= 0.04
                for entry in entries[:5]:
                    ax.text(0.15, y, f"  – {entry['title']}", fontsize=9,
                            color="#555555", va="top")
                    y -= 0.025
                if len(entries) > 5:
                    ax.text(0.15, y, f"  ... and {len(entries) - 5} more",
                            fontsize=9, color="#999999", va="top")
                    y -= 0.025
                y -= 0.015
        pdf.savefig(fig)
        plt.close(fig)

        # Figure pages
        page_count = 0
        for cat in category_order:
            if cat not in categories:
                continue
            entries = categories[cat]

            # Section divider page
            cat_title = cat.replace("_", " ").title()
            fig, ax = plt.subplots(figsize=(11, 8.5))
            ax.axis("off")
            ax.text(0.5, 0.55, cat_title, fontsize=36, fontweight="bold",
                    ha="center", va="center", color="#2e86c1")
            ax.text(0.5, 0.42, f"{len(entries)} figures",
                    fontsize=18, ha="center", va="center", color="#7f8c8d")
            pdf.savefig(fig)
            plt.close(fig)

            for entry in entries:
                fname = entry["filename"]
                fpath = fig_dir / fname
                # Molecules are in a subdirectory
                if not fpath.exists():
                    fpath = fig_dir / "molecules" / fname
                if not fpath.exists():
                    continue

                try:
                    img = plt.imread(str(fpath))
                    fig, ax = plt.subplots(figsize=(11, 8.5))
                    ax.imshow(img)
                    ax.axis("off")
                    ax.set_title(entry["title"], fontsize=12,
                                 fontweight="bold", pad=10)
                    if entry.get("description"):
                        fig.text(0.5, 0.02, entry["description"],
                                 fontsize=8, ha="center", color="#666666")
                    pdf.savefig(fig)
                    plt.close(fig)
                    page_count += 1
                except Exception:
                    plt.close("all")
                    continue

    _log(f"  PDF atlas: {page_count} figure pages -> {output_path}")
    return str(output_path)


# ═══════════════════════════════════════════════════════════════════════
# PHASE C2: XLSX FIGURE EMBEDDING
# ═══════════════════════════════════════════════════════════════════════

def embed_figures_in_xlsx(existing_xlsx: Path, fig_dir: Path,
                          manifest: List[dict],
                          output_xlsx: Path) -> str:
    """Add figure sheets to the existing 5-prong XLSX workbook."""
    _log("Phase C2: Embedding figures into XLSX workbook")

    wb = openpyxl.load_workbook(str(existing_xlsx))

    categories = defaultdict(list)
    for entry in manifest:
        categories[entry["category"]].append(entry)

    category_order = ["thermo", "crystal", "metals", "heating",
                      "alloy", "sampling", "golden", "molecule"]

    sheets_added = 0
    for cat in category_order:
        if cat not in categories:
            continue
        entries = categories[cat]
        cat_title = f"Fig_{cat.title()}"
        # Truncate sheet name to 31 chars (Excel limit)
        cat_title = cat_title[:31]

        ws = wb.create_sheet(cat_title)
        ws.cell(row=1, column=1, value="Figure").font = HEADER_FONT
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=2, value="Title").font = HEADER_FONT
        ws.cell(row=1, column=2).fill = HEADER_FILL
        ws.cell(row=1, column=3, value="Description").font = HEADER_FONT
        ws.cell(row=1, column=3).fill = HEADER_FILL
        ws.cell(row=1, column=4, value="Data Source").font = HEADER_FONT
        ws.cell(row=1, column=4).fill = HEADER_FILL

        row_cursor = 2
        embedded = 0
        for entry in entries:
            fname = entry["filename"]
            fpath = fig_dir / fname
            if not fpath.exists():
                fpath = fig_dir / "molecules" / fname
            if not fpath.exists():
                continue

            ws.cell(row=row_cursor, column=1, value=fname)
            ws.cell(row=row_cursor, column=2, value=entry["title"])
            ws.cell(row=row_cursor, column=3,
                    value=entry.get("description", ""))
            ws.cell(row=row_cursor, column=4,
                    value=entry.get("data_source", ""))

            try:
                img = XlImage(str(fpath))
                # Scale to fit reasonable cell area
                img.width = 480
                img.height = 340
                anchor_cell = f"E{row_cursor}"
                ws.add_image(img, anchor_cell)
                embedded += 1
            except Exception:
                pass

            row_cursor += 20  # space for image height

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 25
        sheets_added += 1
        _log(f"  Sheet '{cat_title}': {embedded} images embedded")

    # Figure index sheet
    ws_idx = wb.create_sheet("Figure Index")
    idx_headers = ["#", "Filename", "Category", "Title",
                   "Description", "Data Source"]
    for ci, h in enumerate(idx_headers, 1):
        cell = ws_idx.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, entry in enumerate(manifest, 1):
        ws_idx.cell(row=i + 1, column=1, value=i)
        ws_idx.cell(row=i + 1, column=2, value=entry["filename"])
        ws_idx.cell(row=i + 1, column=3, value=entry["category"])
        ws_idx.cell(row=i + 1, column=4, value=entry["title"])
        ws_idx.cell(row=i + 1, column=5, value=entry.get("description", ""))
        ws_idx.cell(row=i + 1, column=6, value=entry.get("data_source", ""))

    for col_w in [(1, 5), (2, 35), (3, 12), (4, 35), (5, 40), (6, 30)]:
        ws_idx.column_dimensions[get_column_letter(col_w[0])].width = col_w[1]

    ws_idx.auto_filter.ref = f"A1:F{len(manifest) + 1}"
    sheets_added += 1

    wb.save(str(output_xlsx))
    _log(f"  XLSX with figures saved: {output_xlsx}")
    _log(f"  Added {sheets_added} new sheets, total: {len(wb.sheetnames)}")
    return str(output_xlsx)


# ═══════════════════════════════════════════════════════════════════════
# PHASE C3: LaTeX BACK-MATTER DOCUMENT
# ═══════════════════════════════════════════════════════════════════════

def generate_latex_report(fig_dir: Path, manifest: List[dict],
                          output_path: Path, run_dir: Path) -> str:
    """Generate a LaTeX document with all figures and data tables."""
    _log("Phase C3: Generating LaTeX back-matter document")

    categories = defaultdict(list)
    for entry in manifest:
        categories[entry["category"]].append(entry)

    category_titles = {
        "thermo": "Thermophysical Atlas",
        "crystal": "Crystal Discovery",
        "metals": "Metals \\& Macros",
        "heating": "Heating Simulations",
        "alloy": "Alloy Property Estimation",
        "sampling": "SmartSampling Analysis",
        "golden": "Golden Project Cross-Pillar Discovery",
        "molecule": "Molecular Structure Renderings",
    }
    category_order = ["thermo", "crystal", "metals", "heating",
                      "alloy", "sampling", "golden", "molecule"]

    lines = [
        r"\documentclass[11pt, a4paper]{article}",
        r"\usepackage[margin=1in]{geometry}",
        r"\usepackage{graphicx}",
        r"\usepackage{float}",
        r"\usepackage{subcaption}",
        r"\usepackage{booktabs}",
        r"\usepackage{longtable}",
        r"\usepackage{hyperref}",
        r"\usepackage{xcolor}",
        r"\usepackage{fancyhdr}",
        r"\usepackage{titlesec}",
        "",
        r"\definecolor{vsblue}{HTML}{2E86C1}",
        r"\definecolor{vsdark}{HTML}{2C3E50}",
        "",
        r"\hypersetup{colorlinks=true, linkcolor=vsblue, urlcolor=vsblue}",
        r"\pagestyle{fancy}",
        r"\fancyhead[L]{\textcolor{vsdark}{VSEPR-SIM 5-Prong Report}}",
        r"\fancyhead[R]{\textcolor{vsdark}{\thepage}}",
        r"\fancyfoot[C]{\footnotesize Anti-black-box: every figure traces to source data}",
        "",
        r"\title{\textcolor{vsblue}{\textbf{VSEPR-SIM 5-Prong Data Generation}} \\[0.5em]",
        r"\large Back-Matter: Figures, Renderings \& Data Tables}",
        r"\author{VSEPR-SIM Automated Report Engine}",
        f"\\date{{{datetime.now().strftime('%B %d, %Y')}}}",
        "",
        r"\begin{document}",
        r"\maketitle",
        r"\tableofcontents",
        r"\clearpage",
        "",
    ]

    # Summary section
    lines.append(r"\section{Report Summary}")
    lines.append(r"\begin{itemize}")
    lines.append(f"  \\item Total figures generated: {len(manifest)}")
    for cat in category_order:
        if cat in categories:
            title = category_titles.get(cat, cat.title())
            lines.append(f"  \\item {title}: {len(categories[cat])} figures")
    lines.append(r"\end{itemize}")
    lines.append(r"\clearpage")
    lines.append("")

    # Figure sections
    for cat in category_order:
        if cat not in categories:
            continue
        entries = categories[cat]
        title = category_titles.get(cat, cat.title())
        lines.append(f"\\section{{{title}}}")
        lines.append(f"{len(entries)} figures generated from "
                      f"\\texttt{{{entries[0].get('data_source', 'pipeline data')}}}.")
        lines.append("")

        # Group into pairs for subfigure layouts
        for i in range(0, len(entries), 2):
            batch = entries[i:i + 2]
            if len(batch) == 2:
                f1 = batch[0]["filename"]
                f2 = batch[1]["filename"]
                # Molecules are in subdirectory
                if batch[0]["category"] == "molecule":
                    f1 = f"molecules/{f1}"
                    f2 = f"molecules/{f2}"
                cap1 = batch[0]["title"].replace("_", r"\_")
                cap2 = batch[1]["title"].replace("_", r"\_")
                overall = f"{title} — {i // 2 + 1}"
                label = f"fig:{cat}_{i}"
                lines.append(latex_subfigure_pair(
                    f"figures/{f1}", cap1,
                    f"figures/{f2}", cap2,
                    overall, label, sub_width=0.48,
                ))
            elif len(batch) == 1:
                f1 = batch[0]["filename"]
                if batch[0]["category"] == "molecule":
                    f1 = f"molecules/{f1}"
                cap = batch[0]["title"].replace("_", r"\_")
                label = f"fig:{cat}_{i}"
                lines.append(latex_figure(
                    f"figures/{f1}", cap, label, width=0.85,
                ))
            lines.append("")

        lines.append(r"\clearpage")
        lines.append("")

    # Data tables appendix
    lines.append(r"\appendix")
    lines.append(r"\section{Data Tables}")
    lines.append("")

    # Metals summary table
    metals_csv = run_dir / "prong3_metals.csv"
    metals = _load_csv(metals_csv)
    if metals:
        headers = ["Symbol", "Name", "E (GPa)", "$\\sigma_y$ (MPa)",
                    "k (W/m·K)", "$T_m$ (K)", "$\\rho$ (g/cm$^3$)"]
        table_rows = []
        for r in metals:
            table_rows.append([
                r.get("symbol", ""),
                r.get("name", ""),
                r.get("E_GPa", ""),
                r.get("Sy_MPa", ""),
                r.get("k_WmK", ""),
                r.get("Tm_K", ""),
                r.get("density_gcc", ""),
            ])
        lines.append(latex_table(headers, table_rows,
                                  caption="Metal Property Summary (37 elements)",
                                  label="tab:metals"))
        lines.append(r"\clearpage")

    # Crystal catalog table
    catalog_csv = run_dir / "prong2_catalog.csv"
    catalog = _load_csv(catalog_csv)
    if catalog:
        headers = ["Formula", "Class", "Bravais", "a (\\AA)", "Density"]
        table_rows = []
        for r in catalog:
            table_rows.append([
                r.get("formula", ""),
                r.get("crystal_class", ""),
                r.get("bravais", ""),
                r.get("a_A", ""),
                r.get("density_gcc", ""),
            ])
        lines.append(latex_table(headers, table_rows,
                                  caption="Crystal Catalog (38 prototype lattices)",
                                  label="tab:crystals"))
        lines.append(r"\clearpage")

    # Figure index table
    lines.append(r"\section{Figure Index}")
    lines.append("")
    lines.append(r"\begin{longtable}{rllp{5cm}l}")
    lines.append(r"\toprule")
    lines.append(r"\textbf{\#} & \textbf{Category} & \textbf{Filename} & "
                  r"\textbf{Title} & \textbf{Source} \\")
    lines.append(r"\midrule")
    lines.append(r"\endhead")
    for i, entry in enumerate(manifest, 1):
        cat = entry["category"].replace("_", r"\_")
        fname = entry["filename"].replace("_", r"\_")
        title = entry["title"].replace("_", r"\_")
        src = entry.get("data_source", "").replace("_", r"\_")
        lines.append(f"{i} & {cat} & \\texttt{{{fname}}} & {title} & {src} \\\\")
    lines.append(r"\bottomrule")
    lines.append(r"\end{longtable}")
    lines.append(r"\clearpage")

    lines.append(r"\end{document}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    _log(f"  LaTeX document: {output_path}")
    return str(output_path)


# ═══════════════════════════════════════════════════════════════════════
# PHASE C4: MARKDOWN FIGURE INDEX
# ═══════════════════════════════════════════════════════════════════════

def generate_markdown_index(fig_dir: Path, manifest: List[dict],
                            output_path: Path) -> str:
    """Generate a Markdown figure index with provenance."""
    _log("Phase C4: Generating Markdown figure index")

    categories = defaultdict(list)
    for entry in manifest:
        categories[entry["category"]].append(entry)

    category_order = ["thermo", "crystal", "metals", "heating",
                      "alloy", "sampling", "golden", "molecule"]

    lines = [
        "# VSEPR-SIM 5-Prong Report — Figure Index",
        "",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"**Total figures:** {len(manifest)}  ",
        "",
        "## Summary",
        "",
        "| Category | Figures |",
        "|----------|---------|",
    ]

    for cat in category_order:
        if cat in categories:
            lines.append(f"| {cat.title()} | {len(categories[cat])} |")
    lines.append("")

    for cat in category_order:
        if cat not in categories:
            continue
        entries = categories[cat]
        lines.append(f"## {cat.replace('_', ' ').title()}")
        lines.append("")
        lines.append("| # | Filename | Title | Source |")
        lines.append("|---|----------|-------|--------|")
        for i, entry in enumerate(entries, 1):
            lines.append(
                f"| {i} | `{entry['filename']}` | "
                f"{entry['title']} | {entry.get('data_source', '')} |"
            )
        lines.append("")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    _log(f"  Markdown index: {output_path}")
    return str(output_path)


# ═══════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="VSEPR-SIM 5-Prong Back-Matter & Figure Generation Engine")
    parser.add_argument("--run-dir", type=str, default=None,
                        help="Path to five_prong_run output directory")
    parser.add_argument("--skip-molecules", action="store_true",
                        help="Skip molecular structure renderings")
    parser.add_argument("--skip-latex", action="store_true",
                        help="Skip LaTeX document generation")
    parser.add_argument("--skip-pdf", action="store_true",
                        help="Skip PDF atlas generation")
    parser.add_argument("--max-molecules", type=int, default=50,
                        help="Maximum number of molecules to render")
    parser.add_argument("--skip-species", action="store_true",
                        help="Skip species record export and figures")
    parser.add_argument("--species-shell", action="store_true",
                        help="Launch interactive species record shell (NIST/JANAF/VSEPR/VSPES)")
    args = parser.parse_args()

    # ── Interactive species shell mode (bypasses batch pipeline) ──
    if args.species_shell:
        project_root = Path(__file__).resolve().parent
        species_dir = (Path(args.run_dir) if args.run_dir
                       else project_root / "out" / "five_prong_run") / "species"
        shell = SpeciesShell(species_dir)
        shell.run()
        return {}

    project_root = Path(__file__).resolve().parent
    if args.run_dir:
        run_dir = Path(args.run_dir)
    else:
        run_dir = project_root / "out" / "five_prong_run"

    fig_dir = run_dir / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)

    _log("=" * 70)
    _log("VSEPR-SIM 5-Prong Back-Matter & Figure Generation Engine")
    _log(f"  Run directory: {run_dir}")
    _log(f"  Figure output: {fig_dir}")
    _log(f"  Timestamp: {TIMESTAMP}")
    _log("=" * 70)

    t_total = time.time()
    all_manifest: List[dict] = []

    # ── Phase A: Scientific figures ──
    engine = FigureEngine(run_dir, fig_dir)
    n_thermo = engine.generate_thermo_figures()
    n_crystal = engine.generate_crystal_figures()
    n_metals = engine.generate_metals_figures()
    n_sampling = engine.generate_sampling_figures()
    n_golden = engine.generate_golden_figures()
    all_manifest.extend(engine.manifest)

    # ── Phase B: Molecular renderings ──
    n_mol = 0
    if not args.skip_molecules:
        xyz_dir = project_root / "examples" / "my_molecules"
        renderer = MoleculeRenderer(xyz_dir, fig_dir,
                                     max_molecules=args.max_molecules)
        n_mol = renderer.render_all()
        all_manifest.extend(renderer.manifest)

    # ── Phase D: Species record export and Cp figures ──
    n_species = 0
    species_dir = run_dir / "species"
    if not args.skip_species:
        _log("Phase D: Species record export (NIST/JANAF/VSEPR/VSPES)")
        species_batch_export(species_dir)
        n_species_figs = species_batch_figures(species_dir / "figures")
        n_species = n_species_figs
        _log(f"  Species records exported to {species_dir}")
        _log(f"  Species Cp figures: {n_species_figs}")

    # ── Phase C1: PDF atlas ──
    pdf_path = None
    if not args.skip_pdf:
        pdf_out = run_dir / f"VSEPR_SIM_figure_atlas_{TIMESTAMP}.pdf"
        pdf_path = build_pdf_atlas(fig_dir, pdf_out, all_manifest)

    # ── Phase C2: XLSX figure embedding ──
    existing_xlsx_files = sorted(run_dir.glob("VSEPR_SIM_5prong_*.xlsx"),
                                  key=lambda p: p.stat().st_mtime, reverse=True)
    xlsx_path = None
    if existing_xlsx_files:
        src_xlsx = existing_xlsx_files[0]
        out_xlsx = run_dir / f"VSEPR_SIM_5prong_with_figures_{TIMESTAMP}.xlsx"
        xlsx_path = embed_figures_in_xlsx(src_xlsx, fig_dir,
                                          all_manifest, out_xlsx)

    # ── Phase C3: LaTeX document ──
    latex_path = None
    if not args.skip_latex:
        latex_out = run_dir / "back_matter_report.tex"
        latex_path = generate_latex_report(fig_dir, all_manifest,
                                            latex_out, run_dir)

    # ── Phase C4: Markdown figure index ──
    md_path = generate_markdown_index(fig_dir, all_manifest,
                                       run_dir / "figure_index.md")

    # ── Save manifest as JSON ──
    manifest_path = run_dir / "figure_manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": TIMESTAMP,
            "total_figures": len(all_manifest),
            "figures": all_manifest,
        }, f, indent=2)

    # ── Summary ──
    elapsed = time.time() - t_total
    _log("=" * 70)
    _log("REPORT ENGINE COMPLETE")
    _log(f"  Total figures: {len(all_manifest)}")
    _log(f"    Thermophysical: {n_thermo}")
    _log(f"    Crystal: {n_crystal}")
    _log(f"    Metals & Alloys: {n_metals}")
    _log(f"    SmartSampling: {n_sampling}")
    _log(f"    Golden Project: {n_golden}")
    _log(f"    Molecular: {n_mol}")
    _log(f"    Species Records: {n_species}")
    if pdf_path:
        _log(f"  PDF Atlas: {pdf_path}")
    if xlsx_path:
        _log(f"  XLSX with figures: {xlsx_path}")
    if latex_path:
        _log(f"  LaTeX document: {latex_path}")
    _log(f"  Markdown index: {md_path}")
    _log(f"  Manifest JSON: {manifest_path}")
    _log(f"  Elapsed: {elapsed:.1f}s")
    _log("=" * 70)

    return {
        "total_figures": len(all_manifest),
        "pdf_atlas": pdf_path,
        "xlsx_with_figures": xlsx_path,
        "latex_report": latex_path,
        "markdown_index": md_path,
        "manifest": str(manifest_path),
        "species_dir": str(species_dir) if not args.skip_species else None,
        "species_figures": n_species,
        "elapsed_s": round(elapsed, 2),
    }


if __name__ == "__main__":
    main()
