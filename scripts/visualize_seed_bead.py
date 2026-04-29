#!/usr/bin/env python3
"""
visualize_seed_bead.py — Seed & Bead Model Visualization for Reports

Reads CSV time-series and snapshot data exported by the SeedBeadStepper
and produces publication-quality figures for inclusion in research reports.

Outputs:
  - Energy convergence plot (gradient fill, rolling window band)
  - RMS force convergence plot (dual gradient fills)
  - Environment state evolution (ρ, C, P₂, η, f) — coloured fills
  - Kernel modulation evolution (g_steric, g_elec, g_disp) — gradient bands
  - XY and XZ bead trajectory projections with colorbars
  - Combined 6-panel summary figure (full colour)
  - Excel (.xlsx) workbook export
  - SolidWorks curve-through-XYZ-points export

Usage:
  python visualize_seed_bead.py [--timeseries FILE] [--snapshots FILE] [--output DIR]
  python visualize_seed_bead.py --demo
  python visualize_seed_bead.py --demo --excel    (also export .xlsx)
  python visualize_seed_bead.py --demo --solidworks (also export SolidWorks files)

Reference: VSEPR-SIM Anti-black-box design — every metric inspectable.
"""

import argparse
import os
import sys

import numpy as np

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.gridspec import GridSpec
    from matplotlib.collections import LineCollection
    from matplotlib.colors import Normalize
    import matplotlib.cm as cm
except ImportError:
    print("ERROR: matplotlib required. pip install matplotlib")
    sys.exit(1)


# ============================================================================
# Data Loading
# ============================================================================

def load_timeseries(path):
    """Load time-series CSV from SnapshotGraphCollector."""
    data = np.genfromtxt(path, delimiter=',', names=True, dtype=float)
    return data


def load_snapshots(path):
    """Load snapshot CSV (step, bead_id, x, y, z, eta, rho)."""
    data = np.genfromtxt(path, delimiter=',', names=True, dtype=float)
    return data


# ============================================================================
# Demo Data Generation
# ============================================================================

def generate_demo_data():
    """Generate synthetic demo data for testing."""
    n_steps = 500
    steps = np.arange(n_steps)

    # Simulated convergence curves
    energy = -50.0 * (1 - np.exp(-steps / 80.0)) + np.random.normal(0, 0.5, n_steps)
    rms_force = 5.0 * np.exp(-steps / 60.0) + 0.01 + np.random.normal(0, 0.02, n_steps)
    rms_force = np.maximum(rms_force, 0.001)
    max_force = rms_force * 2.5
    avg_eta = 0.6 * (1 - np.exp(-steps / 120.0))
    avg_rho = 3.0 + 2.0 * (1 - np.exp(-steps / 100.0)) + np.random.normal(0, 0.1, n_steps)
    avg_C = 6.0 + 3.0 * (1 - np.exp(-steps / 90.0))
    avg_P2 = 0.4 * (1 - np.exp(-steps / 150.0))
    avg_target_f = 0.5 * (1 - np.exp(-steps / 100.0))
    max_delta_eta = 0.05 * np.exp(-steps / 80.0)
    dt = 1.0 + 4.0 * (1 - np.exp(-steps / 50.0))
    g_steric = 1.0 + 0.2 * avg_eta
    g_elec = 1.0 - 0.1 * avg_eta
    g_disp = 1.0 + 0.5 * avg_eta
    ke = 2.0 * np.exp(-steps / 40.0)

    # Build structured array
    dtype = np.dtype([
        ('step', float), ('energy', float), ('rms_force', float),
        ('max_force', float), ('avg_eta', float), ('avg_rho', float),
        ('avg_C', float), ('avg_P2', float), ('avg_target_f', float),
        ('max_delta_eta', float), ('dt', float), ('g_steric', float),
        ('g_elec', float), ('g_disp', float), ('ke', float)
    ])
    ts = np.zeros(n_steps, dtype=dtype)
    ts['step'] = steps
    ts['energy'] = energy
    ts['rms_force'] = rms_force
    ts['max_force'] = max_force
    ts['avg_eta'] = avg_eta
    ts['avg_rho'] = avg_rho
    ts['avg_C'] = avg_C
    ts['avg_P2'] = avg_P2
    ts['avg_target_f'] = avg_target_f
    ts['max_delta_eta'] = max_delta_eta
    ts['dt'] = dt
    ts['g_steric'] = g_steric
    ts['g_elec'] = g_elec
    ts['g_disp'] = g_disp
    ts['ke'] = ke

    # Snapshot data (32 beads at 10 time points)
    snap_rows = []
    for t in range(0, n_steps, 50):
        for b in range(32):
            theta = 2 * np.pi * b / 32 + t * 0.001
            r = 5.0 + 2.0 * np.sin(b * 0.5) * (1 - np.exp(-t / 100.0))
            x = r * np.cos(theta)
            y = r * np.sin(theta) + np.random.normal(0, 0.2)
            z = 3.0 * np.sin(b * 0.3 + t * 0.002)
            eta = 0.6 * (1 - np.exp(-t / 120.0)) * (0.8 + 0.2 * np.sin(b))
            rho = 3.0 + np.sin(b * 0.7) + t * 0.005
            snap_rows.append((t, b, x, y, z, eta, rho))

    snap_dtype = np.dtype([
        ('step', float), ('bead_id', float), ('x', float), ('y', float),
        ('z', float), ('eta', float), ('rho', float)
    ])
    snaps = np.array(snap_rows, dtype=snap_dtype)
    return ts, snaps


# ============================================================================
# Plot Functions
# ============================================================================

def _rolling_window(arr, window):
    """Simple rolling mean for confidence bands."""
    if len(arr) < window:
        return arr
    kernel = np.ones(window) / window
    return np.convolve(arr, kernel, mode='same')


def _gradient_line(ax, x, y, cmap_name='coolwarm', linewidth=1.5, alpha=0.9):
    """Draw a line with colour gradient mapped to its own y-value."""
    points = np.column_stack([x, y]).reshape(-1, 1, 2)
    segments = np.concatenate([points[:-1], points[1:]], axis=1)
    norm = Normalize(vmin=np.nanmin(y), vmax=np.nanmax(y))
    lc = LineCollection(segments, cmap=cmap_name, norm=norm)
    lc.set_array(y[:-1])
    lc.set_linewidth(linewidth)
    lc.set_alpha(alpha)
    ax.add_collection(lc)
    ax.set_xlim(x.min(), x.max())
    y_margin = (np.nanmax(y) - np.nanmin(y)) * 0.08 or 1.0
    ax.set_ylim(np.nanmin(y) - y_margin, np.nanmax(y) + y_margin)
    return lc


def plot_energy_convergence(ts, output_dir):
    """Plot energy vs step with gradient line and rolling-mean band."""
    fig, ax = plt.subplots(figsize=(8, 5))
    steps = ts['step'].astype(float)
    energy = ts['energy'].astype(float)

    # Rolling mean ± std band
    win = max(len(energy) // 20, 3)
    e_smooth = _rolling_window(energy, win)
    e_std = np.array([np.std(energy[max(0, i - win):i + win])
                      for i in range(len(energy))])
    ax.fill_between(steps, e_smooth - e_std, e_smooth + e_std,
                    color='#4C72B0', alpha=0.15, label='±1σ band')

    # Gradient-coloured main line
    lc = _gradient_line(ax, steps, energy, cmap_name='winter', linewidth=1.2)
    cbar = fig.colorbar(lc, ax=ax, pad=0.02, aspect=30)
    cbar.set_label('Energy (kcal/mol)', fontsize=10)

    # Smoothed trend overlay
    ax.plot(steps, e_smooth, '--', color='#D65F5F', linewidth=0.9,
            alpha=0.7, label='Rolling mean')

    ax.set_xlabel('Step', fontsize=12)
    ax.set_ylabel('Energy (kcal/mol)', fontsize=12)
    ax.set_title('Seed & Bead — Energy Convergence', fontsize=14)
    ax.legend(fontsize=9, loc='upper right')
    ax.grid(True, alpha=0.25, linestyle=':')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, 'sb_energy_convergence.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_energy_convergence.png")


def plot_force_convergence(ts, output_dir):
    """Plot RMS and max force vs step with gradient fills."""
    fig, ax = plt.subplots(figsize=(8, 5))
    steps = ts['step'].astype(float)
    rms_f = ts['rms_force'].astype(float)
    max_f = ts['max_force'].astype(float)

    # Fill between max and RMS to show gap
    ax.fill_between(steps, rms_f, max_f, alpha=0.12, color='#E24A33',
                    label='Force envelope')

    # RMS force — gradient red-to-green (high to low)
    ax.semilogy(steps, rms_f, '-', color='#E24A33', linewidth=1.2,
                alpha=0.9, label='RMS Force')
    ax.semilogy(steps, max_f, '-', color='#FBC15E', linewidth=0.9,
                alpha=0.75, label='Max Force')

    # Fill under RMS to baseline
    ax.fill_between(steps, rms_f, rms_f.min() * 0.5, alpha=0.06,
                    color='#348ABD')

    ax.set_xlabel('Step', fontsize=12)
    ax.set_ylabel('Force (kcal/(mol·Å))', fontsize=12)
    ax.set_title('Seed & Bead — Force Convergence', fontsize=14)
    ax.legend(fontsize=9, loc='upper right')
    ax.grid(True, alpha=0.25, linestyle=':')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, 'sb_force_convergence.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_force_convergence.png")


def plot_environment_state(ts, output_dir):
    """Plot environment-responsive observables with colour fills."""
    fig, axes = plt.subplots(2, 3, figsize=(15, 9))
    steps = ts['step'].astype(float)

    # Colour palette — rich, distinguishable
    palette = [
        ('#348ABD', '#348ABD'),   # ρ — steel blue
        ('#188A8D', '#188A8D'),   # C — teal
        ('#8B45A6', '#8B45A6'),   # P₂ — purple
        ('#E24A33', '#E24A33'),   # η — warm red
        ('#E5AE38', '#E5AE38'),   # f — golden
        ('#555555', '#555555'),   # max|Δη| — charcoal
    ]
    fields = ['avg_rho', 'avg_C', 'avg_P2', 'avg_eta', 'avg_target_f', 'max_delta_eta']
    titles = ['⟨ρ⟩ (density)', '⟨C⟩ (coordination)', '⟨P₂⟩ (orient. order)',
              '⟨η⟩ (slow state)', '⟨f⟩ (target function)', 'max |Δη| (convergence)']

    for idx, ax in enumerate(axes.flat):
        y = ts[fields[idx]].astype(float)
        clr = palette[idx][0]

        if idx == 5:  # semilogy for max|Δη|
            ax.semilogy(steps, y, '-', color=clr, linewidth=1.0, alpha=0.9)
            y_fill_lo = np.full_like(y, max(y.min() * 0.1, 1e-10))
            ax.fill_between(steps, y_fill_lo, y, alpha=0.10, color=clr)
        else:
            ax.plot(steps, y, '-', color=clr, linewidth=1.1, alpha=0.9)
            # Fill from baseline to curve
            y_base = np.full_like(y, y.min() - (y.max() - y.min()) * 0.05)
            ax.fill_between(steps, y_base, y, alpha=0.12, color=clr)

            # Rolling mean overlay
            win = max(len(y) // 25, 3)
            y_smooth = _rolling_window(y, win)
            ax.plot(steps, y_smooth, '--', color=clr, linewidth=0.7, alpha=0.5)

        if idx == 3:  # η has known bounds
            ax.set_ylim(-0.05, 1.05)

        ax.set_title(titles[idx], fontsize=11, fontweight='bold')
        ax.set_xlabel('Step', fontsize=9)
        ax.grid(True, alpha=0.20, linestyle=':')
        ax.set_facecolor('#FAFAFA')

    fig.suptitle('Seed & Bead — Environment-Responsive State Evolution',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(os.path.join(output_dir, 'sb_environment_state.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_environment_state.png")


def plot_kernel_modulation(ts, output_dir):
    """Plot kernel modulation factors with gradient bands."""
    fig, ax = plt.subplots(figsize=(8, 5))
    steps = ts['step'].astype(float)

    channels = [
        ('g_steric', '#E24A33', 'g_steric (hardening)'),
        ('g_elec',   '#348ABD', 'g_elec (screening)'),
        ('g_disp',   '#2CA02C', 'g_disp (enhancement)'),
    ]

    # Fill bands between 1.0 baseline and each channel
    for field, clr, lbl in channels:
        y = ts[field].astype(float)
        ax.fill_between(steps, 1.0, y, alpha=0.10, color=clr)
        ax.plot(steps, y, '-', color=clr, linewidth=1.3, alpha=0.9, label=lbl)

    ax.axhline(y=1.0, color='#555555', linestyle='--', linewidth=1.0,
               alpha=0.6, label='Unmodulated')

    ax.set_xlabel('Step', fontsize=12)
    ax.set_ylabel('Modulation Factor g', fontsize=12)
    ax.set_title('Seed & Bead — Kernel Modulation K_k · (1 + γ_k · η̄)', fontsize=14)
    ax.legend(fontsize=9, loc='best', framealpha=0.85)
    ax.grid(True, alpha=0.20, linestyle=':')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, 'sb_kernel_modulation.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_kernel_modulation.png")


def plot_trajectory_xy(snaps, output_dir):
    """XY plane projection with time-gradient trails and η colourbar."""
    fig, ax = plt.subplots(figsize=(9, 8))

    n_beads = int(np.max(snaps['bead_id'])) + 1
    trail_cmap = plt.cm.viridis
    eta_cmap = plt.cm.RdYlBu_r

    # Collect final η for colourbar
    final_etas = []
    final_xy = []

    for bid in range(min(n_beads, 32)):
        mask = snaps['bead_id'] == bid
        bead_data = snaps[mask]
        if len(bead_data) == 0:
            continue

        x = bead_data['x'].astype(float)
        y = bead_data['y'].astype(float)
        t = np.linspace(0, 1, len(x))

        # Time-gradient trail using LineCollection
        points = np.column_stack([x, y]).reshape(-1, 1, 2)
        segments = np.concatenate([points[:-1], points[1:]], axis=1)
        lc = LineCollection(segments, cmap=trail_cmap,
                            norm=Normalize(0, 1))
        lc.set_array(t[:-1])
        lc.set_linewidth(0.8)
        lc.set_alpha(0.55)
        ax.add_collection(lc)

        final_etas.append(bead_data['eta'][-1])
        final_xy.append((x[-1], y[-1]))

    # Final positions scatter with η colouring
    if final_xy:
        fx = np.array([p[0] for p in final_xy])
        fy = np.array([p[1] for p in final_xy])
        fe = np.array(final_etas)
        sc = ax.scatter(fx, fy, c=fe, cmap=eta_cmap, vmin=0, vmax=1,
                        s=70, edgecolors='black', linewidth=0.6, zorder=5)
        cbar = fig.colorbar(sc, ax=ax, pad=0.02, aspect=30, shrink=0.85)
        cbar.set_label('Final η', fontsize=11)

    ax.autoscale_view()
    ax.set_xlabel('X (Å)', fontsize=12)
    ax.set_ylabel('Y (Å)', fontsize=12)
    ax.set_title('Seed & Bead — XY Trajectory Projection', fontsize=14)
    ax.set_aspect('equal')
    ax.grid(True, alpha=0.20, linestyle=':')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, 'sb_trajectory_xy.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_trajectory_xy.png")


def plot_trajectory_xz(snaps, output_dir):
    """XZ plane projection with time-gradient trails and ρ colourbar."""
    fig, ax = plt.subplots(figsize=(9, 8))

    n_beads = int(np.max(snaps['bead_id'])) + 1
    trail_cmap = plt.cm.plasma

    final_rhos = []
    final_xz = []

    for bid in range(min(n_beads, 32)):
        mask = snaps['bead_id'] == bid
        bead_data = snaps[mask]
        if len(bead_data) == 0:
            continue

        x = bead_data['x'].astype(float)
        z = bead_data['z'].astype(float)
        t = np.linspace(0, 1, len(x))

        # Time-gradient trail
        points = np.column_stack([x, z]).reshape(-1, 1, 2)
        segments = np.concatenate([points[:-1], points[1:]], axis=1)
        lc = LineCollection(segments, cmap=trail_cmap,
                            norm=Normalize(0, 1))
        lc.set_array(t[:-1])
        lc.set_linewidth(0.8)
        lc.set_alpha(0.55)
        ax.add_collection(lc)

        final_rhos.append(bead_data['rho'][-1])
        final_xz.append((x[-1], z[-1]))

    # Final positions scatter with ρ colouring
    if final_xz:
        fx = np.array([p[0] for p in final_xz])
        fz = np.array([p[1] for p in final_xz])
        fr = np.array(final_rhos)
        sc = ax.scatter(fx, fz, c=fr, cmap='inferno', vmin=0,
                        vmax=max(fr.max(), 1.0),
                        s=70, edgecolors='black', linewidth=0.6, zorder=5)
        cbar = fig.colorbar(sc, ax=ax, pad=0.02, aspect=30, shrink=0.85)
        cbar.set_label('Final ρ', fontsize=11)

    ax.autoscale_view()
    ax.set_xlabel('X (Å)', fontsize=12)
    ax.set_ylabel('Z (Å)', fontsize=12)
    ax.set_title('Seed & Bead — XZ Trajectory Projection', fontsize=14)
    ax.set_aspect('equal')
    ax.grid(True, alpha=0.20, linestyle=':')
    ax.set_facecolor('#FAFAFA')
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, 'sb_trajectory_xz.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_trajectory_xz.png")


def plot_combined_report(ts, snaps, output_dir):
    """6-panel combined summary with full colour, fills, and colorbars."""
    fig = plt.figure(figsize=(19, 13))
    gs = GridSpec(2, 3, figure=fig, hspace=0.38, wspace=0.35)
    steps = ts['step'].astype(float)

    # Panel 1: Energy — gradient fill
    ax1 = fig.add_subplot(gs[0, 0])
    energy = ts['energy'].astype(float)
    ax1.plot(steps, energy, '-', color='#348ABD', linewidth=1.0, alpha=0.9)
    ax1.fill_between(steps, energy.min(), energy, alpha=0.12, color='#348ABD')
    win = max(len(energy) // 25, 3)
    ax1.plot(steps, _rolling_window(energy, win), '--', color='#D65F5F',
             linewidth=0.7, alpha=0.6)
    ax1.set_xlabel('Step', fontsize=9)
    ax1.set_ylabel('Energy (kcal/mol)', fontsize=9)
    ax1.set_title('(a) Energy Convergence', fontsize=11, fontweight='bold')
    ax1.grid(True, alpha=0.20, linestyle=':')
    ax1.set_facecolor('#FAFAFA')

    # Panel 2: Force — envelope fill
    ax2 = fig.add_subplot(gs[0, 1])
    rms_f = ts['rms_force'].astype(float)
    max_f = ts['max_force'].astype(float)
    ax2.fill_between(steps, rms_f, max_f, alpha=0.12, color='#E24A33')
    ax2.semilogy(steps, rms_f, '-', color='#E24A33', linewidth=1.0, label='RMS')
    ax2.semilogy(steps, max_f, '-', color='#FBC15E', linewidth=0.8,
                 alpha=0.7, label='Max')
    ax2.set_xlabel('Step', fontsize=9)
    ax2.set_ylabel('Force', fontsize=9)
    ax2.set_title('(b) Force Convergence', fontsize=11, fontweight='bold')
    ax2.legend(fontsize=8, loc='upper right')
    ax2.grid(True, alpha=0.20, linestyle=':')
    ax2.set_facecolor('#FAFAFA')

    # Panel 3: η + f — dual colour fill
    ax3 = fig.add_subplot(gs[0, 2])
    eta = ts['avg_eta'].astype(float)
    tgt = ts['avg_target_f'].astype(float)
    ax3.fill_between(steps, 0, eta, alpha=0.12, color='#E24A33')
    ax3.fill_between(steps, 0, tgt, alpha=0.10, color='#E5AE38')
    ax3.plot(steps, eta, '-', color='#E24A33', linewidth=1.1, label='⟨η⟩')
    ax3.plot(steps, tgt, '-', color='#E5AE38', linewidth=0.9, label='⟨f⟩')
    ax3.set_xlabel('Step', fontsize=9)
    ax3.set_ylabel('Value', fontsize=9)
    ax3.set_title('(c) Slow State & Target', fontsize=11, fontweight='bold')
    ax3.set_ylim(-0.05, 1.05)
    ax3.legend(fontsize=8, loc='lower right')
    ax3.grid(True, alpha=0.20, linestyle=':')
    ax3.set_facecolor('#FAFAFA')

    # Panel 4: Kernel modulation — fill bands
    ax4 = fig.add_subplot(gs[1, 0])
    for field, clr, lbl in [('g_steric', '#E24A33', 'steric'),
                             ('g_elec', '#348ABD', 'elec'),
                             ('g_disp', '#2CA02C', 'disp')]:
        y = ts[field].astype(float)
        ax4.fill_between(steps, 1.0, y, alpha=0.10, color=clr)
        ax4.plot(steps, y, '-', color=clr, linewidth=1.0, label=lbl)
    ax4.axhline(y=1.0, color='#555555', linestyle='--', linewidth=0.8, alpha=0.5)
    ax4.set_xlabel('Step', fontsize=9)
    ax4.set_ylabel('g', fontsize=9)
    ax4.set_title('(d) Kernel Modulation', fontsize=11, fontweight='bold')
    ax4.legend(fontsize=8, loc='best')
    ax4.grid(True, alpha=0.20, linestyle=':')
    ax4.set_facecolor('#FAFAFA')

    # Panel 5: XY trajectory — colour trails + η scatter
    ax5 = fig.add_subplot(gs[1, 1])
    if snaps is not None and len(snaps) > 0:
        n_beads = int(np.max(snaps['bead_id'])) + 1
        trail_cm = plt.cm.viridis
        final_xy, final_e = [], []
        for bid in range(min(n_beads, 32)):
            mask = snaps['bead_id'] == bid
            bd = snaps[mask]
            if len(bd) == 0:
                continue
            x, y = bd['x'].astype(float), bd['y'].astype(float)
            t = np.linspace(0, 1, len(x))
            pts = np.column_stack([x, y]).reshape(-1, 1, 2)
            segs = np.concatenate([pts[:-1], pts[1:]], axis=1)
            lc = LineCollection(segs, cmap=trail_cm, norm=Normalize(0, 1))
            lc.set_array(t[:-1])
            lc.set_linewidth(0.6)
            lc.set_alpha(0.5)
            ax5.add_collection(lc)
            final_xy.append((x[-1], y[-1]))
            final_e.append(bd['eta'][-1])
        if final_xy:
            fx = np.array([p[0] for p in final_xy])
            fy = np.array([p[1] for p in final_xy])
            fe = np.array(final_e)
            sc = ax5.scatter(fx, fy, c=fe, cmap='RdYlBu_r', vmin=0, vmax=1,
                             s=30, edgecolors='black', linewidth=0.4, zorder=5)
            fig.colorbar(sc, ax=ax5, pad=0.02, aspect=25, shrink=0.8,
                         label='η')
        ax5.autoscale_view()
    ax5.set_xlabel('X (Å)', fontsize=9)
    ax5.set_ylabel('Y (Å)', fontsize=9)
    ax5.set_title('(e) XY Trajectories', fontsize=11, fontweight='bold')
    ax5.set_aspect('equal')
    ax5.grid(True, alpha=0.20, linestyle=':')
    ax5.set_facecolor('#FAFAFA')

    # Panel 6: XZ trajectory — colour trails + ρ scatter
    ax6 = fig.add_subplot(gs[1, 2])
    if snaps is not None and len(snaps) > 0:
        final_xz, final_r = [], []
        for bid in range(min(n_beads, 32)):
            mask = snaps['bead_id'] == bid
            bd = snaps[mask]
            if len(bd) == 0:
                continue
            x, z = bd['x'].astype(float), bd['z'].astype(float)
            t = np.linspace(0, 1, len(x))
            pts = np.column_stack([x, z]).reshape(-1, 1, 2)
            segs = np.concatenate([pts[:-1], pts[1:]], axis=1)
            lc = LineCollection(segs, cmap='plasma', norm=Normalize(0, 1))
            lc.set_array(t[:-1])
            lc.set_linewidth(0.6)
            lc.set_alpha(0.5)
            ax6.add_collection(lc)
            final_xz.append((x[-1], z[-1]))
            final_r.append(bd['rho'][-1])
        if final_xz:
            fx = np.array([p[0] for p in final_xz])
            fz = np.array([p[1] for p in final_xz])
            fr = np.array(final_r)
            sc = ax6.scatter(fx, fz, c=fr, cmap='inferno', vmin=0,
                             vmax=max(fr.max(), 1.0),
                             s=30, edgecolors='black', linewidth=0.4, zorder=5)
            fig.colorbar(sc, ax=ax6, pad=0.02, aspect=25, shrink=0.8,
                         label='ρ')
        ax6.autoscale_view()
    ax6.set_xlabel('X (Å)', fontsize=9)
    ax6.set_ylabel('Z (Å)', fontsize=9)
    ax6.set_title('(f) XZ Trajectories', fontsize=11, fontweight='bold')
    ax6.set_aspect('equal')
    ax6.grid(True, alpha=0.20, linestyle=':')
    ax6.set_facecolor('#FAFAFA')

    fig.suptitle('VSEPR-SIM: Seed & Bead Model — 6+9 Steady-State Summary',
                 fontsize=16, fontweight='bold')
    fig.savefig(os.path.join(output_dir, 'sb_combined_report.png'), dpi=150)
    plt.close(fig)
    print(f"  Written: sb_combined_report.png")


# ============================================================================
# Excel Export (.xlsx)
# ============================================================================

def export_excel(ts, snaps, output_dir):
    """Export time-series and snapshot data to an Excel workbook (.xlsx).

    Requires openpyxl.  Each data channel gets its own sheet with
    column headers, units, and formatted numeric cells.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, numbers
    except ImportError:
        print("  WARN: openpyxl not installed — skipping Excel export.")
        print("        pip install openpyxl")
        return

    wb = Workbook()

    # ── Sheet 1: Timeseries ────────────────────────────────────────────────
    ws_ts = wb.active
    ws_ts.title = 'Timeseries'

    ts_cols = [
        ('Step', 'step', '#'),
        ('Energy', 'energy', 'kcal/mol'),
        ('RMS Force', 'rms_force', 'kcal/(mol·Å)'),
        ('Max Force', 'max_force', 'kcal/(mol·Å)'),
        ('Mean η', 'avg_eta', '—'),
        ('Mean ρ', 'avg_rho', '—'),
        ('Mean C', 'avg_C', '—'),
        ('Mean P₂', 'avg_P2', '—'),
        ('Mean f', 'avg_target_f', '—'),
        ('Max |Δη|', 'max_delta_eta', '—'),
        ('dt', 'dt', 'fs'),
        ('g_steric', 'g_steric', '—'),
        ('g_elec', 'g_elec', '—'),
        ('g_disp', 'g_disp', '—'),
        ('KE', 'ke', 'kcal/mol'),
    ]

    header_font = Font(bold=True, size=11)
    unit_font = Font(italic=True, size=10, color='666666')
    header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0',
                              fill_type='solid')

    # Headers
    for c, (name, _, unit) in enumerate(ts_cols, 1):
        cell = ws_ts.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        # Unit row
        ucell = ws_ts.cell(row=2, column=c, value=f'[{unit}]')
        ucell.font = unit_font
        ucell.alignment = Alignment(horizontal='center')

    # Data rows
    for i in range(len(ts)):
        for c, (_, field, _) in enumerate(ts_cols, 1):
            cell = ws_ts.cell(row=i + 3, column=c, value=float(ts[field][i]))
            cell.number_format = '0.000000' if c > 1 else '0'

    # Auto-width
    for c in range(1, len(ts_cols) + 1):
        ws_ts.column_dimensions[
            ws_ts.cell(row=1, column=c).column_letter
        ].width = 15

    # ── Sheet 2: Snapshots ─────────────────────────────────────────────────
    if snaps is not None and len(snaps) > 0:
        ws_sn = wb.create_sheet('Snapshots')
        snap_cols = [
            ('Step', 'step', '#'),
            ('Bead ID', 'bead_id', '#'),
            ('X', 'x', 'Å'),
            ('Y', 'y', 'Å'),
            ('Z', 'z', 'Å'),
            ('η', 'eta', '—'),
            ('ρ', 'rho', '—'),
        ]

        for c, (name, _, unit) in enumerate(snap_cols, 1):
            cell = ws_sn.cell(row=1, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ucell = ws_sn.cell(row=2, column=c, value=f'[{unit}]')
            ucell.font = unit_font
            ucell.alignment = Alignment(horizontal='center')

        for i in range(len(snaps)):
            for c, (_, field, _) in enumerate(snap_cols, 1):
                cell = ws_sn.cell(row=i + 3, column=c,
                                  value=float(snaps[field][i]))
                cell.number_format = '0.000000' if c > 2 else '0'

        for c in range(1, len(snap_cols) + 1):
            ws_sn.column_dimensions[
                ws_sn.cell(row=1, column=c).column_letter
            ].width = 14

    # ── Sheet 3: Summary ───────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    summary_rows = [
        ('Metric', 'Value', 'Unit'),
        ('Final Energy', float(ts['energy'][-1]), 'kcal/mol'),
        ('Final RMS Force', float(ts['rms_force'][-1]), 'kcal/(mol·Å)'),
        ('Final Mean η', float(ts['avg_eta'][-1]), '—'),
        ('Final Mean ρ', float(ts['avg_rho'][-1]), '—'),
        ('Final g_steric', float(ts['g_steric'][-1]), '—'),
        ('Final g_elec', float(ts['g_elec'][-1]), '—'),
        ('Final g_disp', float(ts['g_disp'][-1]), '—'),
        ('Total Steps', int(ts['step'][-1]), '#'),
    ]
    for r, row_data in enumerate(summary_rows, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill

    for c in range(1, 4):
        ws_sum.column_dimensions[
            ws_sum.cell(row=1, column=c).column_letter
        ].width = 20

    xlsx_path = os.path.join(output_dir, 'sb_seed_bead_data.xlsx')
    wb.save(xlsx_path)
    print(f"  Written: sb_seed_bead_data.xlsx")


# ============================================================================
# SolidWorks Export (Curve-Through-XYZ-Points)
# ============================================================================

def export_solidworks(snaps, output_dir):
    """Export bead positions in SolidWorks-compatible formats.

    Produces:
      1. .sldcrv  — SolidWorks "Curve Through XYZ Points" format
                    (tab-delimited X Y Z, one point per line)
      2. .xyz     — Standard XYZ point cloud (atom-style, for import)
      3. .csv     — Plain X,Y,Z coordinate CSV for SolidWorks "Insert > Curve"

    SolidWorks import path:
      Insert > Curve > Curve Through XYZ Points > Browse > select .sldcrv
    """
    if snaps is None or len(snaps) == 0:
        print("  WARN: No snapshot data — skipping SolidWorks export.")
        return

    # Use final timestep only for the structural export
    final_step = snaps['step'].max()
    mask = snaps['step'] == final_step
    final = snaps[mask]

    if len(final) == 0:
        print("  WARN: No beads at final step — skipping SolidWorks export.")
        return

    # ── 1. SolidWorks Curve-Through-XYZ-Points (.sldcrv) ──────────────────
    sldcrv_path = os.path.join(output_dir, 'sb_final_structure.sldcrv')
    with open(sldcrv_path, 'w') as f:
        for i in range(len(final)):
            f.write(f"{final['x'][i]:.6f}\t"
                    f"{final['y'][i]:.6f}\t"
                    f"{final['z'][i]:.6f}\n")
    print(f"  Written: sb_final_structure.sldcrv")

    # ── 2. XYZ point cloud (.xyz) ─────────────────────────────────────────
    xyz_path = os.path.join(output_dir, 'sb_final_structure.xyz')
    with open(xyz_path, 'w') as f:
        f.write(f"{len(final)}\n")
        f.write(f"VSEPR-SIM Seed & Bead — final structure at step {int(final_step)}\n")
        for i in range(len(final)):
            f.write(f"Bead  {final['x'][i]:12.6f}  "
                    f"{final['y'][i]:12.6f}  "
                    f"{final['z'][i]:12.6f}\n")
    print(f"  Written: sb_final_structure.xyz")

    # ── 3. SolidWorks-importable CSV ──────────────────────────────────────
    csv_path = os.path.join(output_dir, 'sb_solidworks_points.csv')
    with open(csv_path, 'w') as f:
        f.write("X,Y,Z,eta,rho,bead_id\n")
        for i in range(len(final)):
            f.write(f"{final['x'][i]:.6f},{final['y'][i]:.6f},"
                    f"{final['z'][i]:.6f},{final['eta'][i]:.6f},"
                    f"{final['rho'][i]:.6f},{int(final['bead_id'][i])}\n")
    print(f"  Written: sb_solidworks_points.csv")

    # ── 4. Full trajectory .sldcrv (all timesteps, per-bead curves) ───────
    traj_dir = os.path.join(output_dir, 'solidworks_trajectories')
    os.makedirs(traj_dir, exist_ok=True)
    n_beads = int(np.max(snaps['bead_id'])) + 1
    for bid in range(min(n_beads, 64)):
        mask = snaps['bead_id'] == bid
        bd = snaps[mask]
        if len(bd) == 0:
            continue
        crv_path = os.path.join(traj_dir, f'bead_{bid:03d}.sldcrv')
        with open(crv_path, 'w') as f:
            for j in range(len(bd)):
                f.write(f"{bd['x'][j]:.6f}\t"
                        f"{bd['y'][j]:.6f}\t"
                        f"{bd['z'][j]:.6f}\n")
    print(f"  Written: solidworks_trajectories/ ({min(n_beads, 64)} bead curves)")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Seed & Bead model visualization for research reports')
    parser.add_argument('--timeseries', type=str, default=None,
                        help='Path to timeseries CSV')
    parser.add_argument('--snapshots', type=str, default=None,
                        help='Path to snapshots CSV')
    parser.add_argument('--output', type=str, default='figures/',
                        help='Output directory for figures')
    parser.add_argument('--demo', action='store_true',
                        help='Generate demo data and plots')
    parser.add_argument('--excel', action='store_true',
                        help='Export data to Excel .xlsx workbook')
    parser.add_argument('--solidworks', action='store_true',
                        help='Export SolidWorks-compatible point/curve files')
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    if args.demo:
        print("Generating demo data...")
        ts, snaps = generate_demo_data()
    else:
        if args.timeseries is None:
            print("ERROR: provide --timeseries or --demo")
            sys.exit(1)
        print(f"Loading timeseries from {args.timeseries}...")
        ts = load_timeseries(args.timeseries)
        snaps = None
        if args.snapshots:
            print(f"Loading snapshots from {args.snapshots}...")
            snaps = load_snapshots(args.snapshots)

    print(f"\nGenerating figures in {args.output}/...")
    plot_energy_convergence(ts, args.output)
    plot_force_convergence(ts, args.output)
    plot_environment_state(ts, args.output)
    plot_kernel_modulation(ts, args.output)

    if snaps is not None and len(snaps) > 0:
        plot_trajectory_xy(snaps, args.output)
        plot_trajectory_xz(snaps, args.output)

    plot_combined_report(ts, snaps, args.output)

    # Excel export
    if args.excel:
        print("\nExporting Excel workbook...")
        export_excel(ts, snaps, args.output)

    # SolidWorks export
    if args.solidworks:
        print("\nExporting SolidWorks data...")
        export_solidworks(snaps, args.output)

    print("\nDone. All outputs generated.")


if __name__ == '__main__':
    main()
