"""
five_prong_data_gen.py -- VSEPR-SIM 5-Prong Large Data Creation Engine
=======================================================================

Prong 1: Thermophysical Atlas  (SteamTables++ full grid sweep, all materials)
Prong 2: Crystal Discovery     (full catalog + supercells + defects + substitutions)
Prong 3: Metals & Macros       (all metals, all missions, alloy ROM, heating sims)
Prong 4: SmartSampling         (adaptive refinement across all domains)
Prong 5: Golden Project        (cross-pillar discovery missions, unified ranking)

Design: accuracy over speed. Parallelism via concurrent.futures for throughput.
Deterministic: seed-controlled throughout.
"""

from __future__ import annotations

import csv
import json
import math
import os
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SCRIPT_DIR
if not (_PROJECT_ROOT / "pykernel").is_dir():
    if (_PROJECT_ROOT.parent / "pykernel").is_dir():
        _PROJECT_ROOT = _PROJECT_ROOT.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from pykernel.pillars.steam_tables import (
    get_all_materials, compute_state, compute_saturation_line,
    StatePoint, MaterialDef, Tier,
)
from pykernel.pillars.crystal_discovery import (
    CrystalCatalog, CrystalDiscoveryEngine, CrystalClass,
    export_catalog_csv,
)
from pykernel.pillars.metals_macros import (
    _METAL_DB, MacroProperties, MissionSpec, MissionObjective,
    AlloyEstimator, run_mission, simulate_heating, score_material,
    export_metals_csv,
)
from pykernel.pillars.smart_sampling import (
    SmartSampler, SamplingWeights, SamplingTier,
    create_steam_sampler, create_crystal_sampler, create_alloy_sampler,
)
from pykernel.pillars.golden_project import (
    DiscoveryAtlas, DiscoveryMission, MissionType,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

OUT_ROOT = _PROJECT_ROOT / "out" / "five_prong_run"
OUT_ROOT.mkdir(parents=True, exist_ok=True)
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
_LOG_FILE = OUT_ROOT / f"run_{TIMESTAMP}.log"


def log(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ===========================================================================
# XLSX helpers
# ===========================================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL_A = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
ALT_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_data_rows(ws, nrows: int, ncols: int):
    for row_idx in range(2, min(nrows + 2, 5002)):
        fill = ALT_FILL_A if row_idx % 2 == 0 else ALT_FILL_B
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, ncols: int, max_width: int = 28):
    for col in range(1, ncols + 1):
        max_len = 0
        letter = get_column_letter(col)
        for cell in list(ws.iter_rows(min_col=col, max_col=col, max_row=min(ws.max_row, 100), values_only=False)):
            for c in cell:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, max_width)


def _add_autofilter(ws, ncols: int):
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    ncols = len(headers)
    nrows = len(rows)
    _style_header(ws, ncols)
    _style_data_rows(ws, nrows, ncols)
    _auto_width(ws, ncols)
    _add_autofilter(ws, ncols)


def _add_line_chart(ws, title: str, x_col: int, y_cols: List[int],
                    row_end: Optional[int] = None):
    if row_end is None:
        row_end = ws.max_row
    if row_end <= 2:
        return
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = 22
    chart.height = 14
    cats = Reference(ws, min_col=x_col, min_row=2, max_row=row_end)
    chart.set_categories(cats)
    for yc in y_cols:
        data = Reference(ws, min_col=yc, min_row=1, max_row=row_end)
        chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, f"A{row_end + 3}")


# ===========================================================================
# PRONG 1 -- Thermophysical Atlas
# ===========================================================================

def _compute_material_grid(formula: str, n_T: int, n_P: int) -> Tuple[str, List[dict]]:
    mat_db = get_all_materials()
    mat = mat_db.get(formula)
    if mat is None:
        return formula, []

    rows = []
    Tc = mat.critical.T_c
    Pc = mat.critical.P_c
    T_lo = max(Tc * 0.30, 100.0)
    T_hi = Tc * 1.60
    P_lo = 1e4
    P_hi = Pc * 2.5

    T_step = (T_hi - T_lo) / max(n_T - 1, 1)
    P_step = (P_hi - P_lo) / max(n_P - 1, 1)

    for i_T in range(n_T):
        T = T_lo + i_T * T_step
        for i_P in range(n_P):
            P = P_lo + i_P * P_step
            try:
                sp = compute_state(formula, T, P)
                rows.append({
                    "material": formula,
                    "T_K": round(T, 3),
                    "P_Pa": round(P, 1),
                    "phase": sp.phase.name if hasattr(sp.phase, "name") else str(sp.phase),
                    "h_Jkg": round(sp.h, 4) if sp.h else None,
                    "s_JkgK": round(sp.s, 6) if sp.s else None,
                    "v_m3kg": f"{sp.v:.6e}" if sp.v else None,
                    "cp_JkgK": round(sp.c_p, 4) if sp.c_p else None,
                    "cv_JkgK": round(sp.c_v, 4) if sp.c_v else None,
                    "mu_Pas": f"{sp.mu:.4e}" if sp.mu else None,
                    "k_WmK": round(sp.k_thermal, 6) if sp.k_thermal else None,
                    "Pr": round(sp.Pr, 4) if sp.Pr else None,
                    "quality_x": round(sp.x, 4) if sp.x is not None else None,
                    "Z_comp": round(sp.Z, 6) if sp.Z else None,
                })
            except Exception:
                pass
    return formula, rows


def _compute_saturation(formula: str, n_points: int) -> Tuple[str, List[dict]]:
    try:
        sat_data = compute_saturation_line(formula, n_points=n_points)
        rows = []
        for sp in sat_data:
            rows.append({
                "material": formula,
                "T_K": round(sp.T, 3),
                "P_Pa": round(sp.P, 1),
                "phase": sp.phase.name if hasattr(sp.phase, "name") else str(sp.phase),
                "h_Jkg": round(sp.h, 4) if sp.h else None,
                "s_JkgK": round(sp.s, 6) if sp.s else None,
                "v_m3kg": f"{sp.v:.6e}" if sp.v else None,
            })
        return formula, rows
    except Exception:
        return formula, []


def prong1_thermo(n_T: int = 80, n_P: int = 80, sat_pts: int = 200) -> Dict[str, Any]:
    log("=" * 70)
    log("PRONG 1: Thermophysical Atlas -- starting")
    log(f"  Grid resolution: {n_T} x {n_P} = {n_T * n_P} state points per material")
    log(f"  Saturation curve: {sat_pts} points per material")

    all_mats = get_all_materials()
    formulas = sorted(all_mats.keys())
    log(f"  Materials registered: {len(formulas)}")

    t0 = time.time()

    grid_results: Dict[str, List[dict]] = {}
    with ThreadPoolExecutor(max_workers=min(12, len(formulas))) as pool:
        futures = {pool.submit(_compute_material_grid, f, n_T, n_P): f for f in formulas}
        for fut in as_completed(futures):
            formula, rows = fut.result()
            grid_results[formula] = rows
            log(f"  [grid] {formula}: {len(rows)} state points")

    sat_results: Dict[str, List[dict]] = {}
    with ThreadPoolExecutor(max_workers=min(12, len(formulas))) as pool:
        futures = {pool.submit(_compute_saturation, f, sat_pts): f for f in formulas}
        for fut in as_completed(futures):
            formula, rows = fut.result()
            sat_results[formula] = rows
            log(f"  [sat]  {formula}: {len(rows)} saturation points")

    grid_csv = OUT_ROOT / "prong1_state_grid.csv"
    total_grid_rows = 0
    with open(grid_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "material", "T_K", "P_Pa", "phase", "h_Jkg", "s_JkgK",
            "v_m3kg", "cp_JkgK", "cv_JkgK", "mu_Pas", "k_WmK",
            "Pr", "quality_x", "Z_comp"])
        w.writeheader()
        for formula in formulas:
            for row in grid_results.get(formula, []):
                w.writerow(row)
                total_grid_rows += 1

    sat_csv = OUT_ROOT / "prong1_saturation.csv"
    total_sat_rows = 0
    with open(sat_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "material", "T_K", "P_Pa", "phase", "h_Jkg", "s_JkgK", "v_m3kg"])
        w.writeheader()
        for formula in formulas:
            for row in sat_results.get(formula, []):
                w.writerow(row)
                total_sat_rows += 1

    elapsed = time.time() - t0
    summary = {
        "prong": "1_thermo",
        "materials": len(formulas),
        "grid_points": total_grid_rows,
        "sat_points": total_sat_rows,
        "elapsed_s": round(elapsed, 2),
    }
    log(f"PRONG 1 complete: {total_grid_rows} grid + {total_sat_rows} sat = "
        f"{total_grid_rows + total_sat_rows} total in {elapsed:.1f}s")
    return {"summary": summary, "grid": grid_results, "sat": sat_results,
            "formulas": formulas}


# ===========================================================================
# PRONG 2 -- Crystal Discovery
# ===========================================================================

def prong2_crystals(seed: int = 42) -> Dict[str, Any]:
    log("=" * 70)
    log("PRONG 2: Crystal Discovery -- starting")
    t0 = time.time()

    catalog = CrystalCatalog()
    engine = CrystalDiscoveryEngine(catalog, seed=seed)

    all_entries = catalog.search()
    log(f"  Catalog entries: {len(all_entries)}")

    supercell_data = []
    supercell_sizes = [(2,2,2), (3,3,3), (4,4,4), (2,2,1), (3,3,1), (2,3,2), (5,5,5)]
    for entry in all_entries:
        for nx, ny, nz in supercell_sizes:
            try:
                sc = catalog.generate_supercell(entry, nx, ny, nz)
                supercell_data.append({
                    "base_formula": entry.formula,
                    "supercell": f"{nx}x{ny}x{nz}",
                    "n_atoms": sc.n_atoms,
                    "volume_A3": round(sc.volume, 3),
                    "lattice": entry.bravais.value if hasattr(entry.bravais, "value") else str(entry.bravais),
                    "crystal_class": entry.crystal_class.value if hasattr(entry.crystal_class, "value") else str(entry.crystal_class),
                    "provenance": sc.provenance_hash,
                })
            except Exception:
                pass
    log(f"  Supercells generated: {len(supercell_data)}")

    defect_data = []
    for entry in all_entries:
        for defect_type in ["vacancy", "interstitial", "substitution"]:
            try:
                defected = catalog.insert_defect(entry, defect_type=defect_type)
                defect_data.append({
                    "base_formula": entry.formula,
                    "defect_type": defect_type,
                    "n_atoms": defected.n_atoms,
                    "volume_A3": round(defected.volume, 3),
                    "provenance": defected.provenance_hash,
                })
            except Exception:
                pass
    log(f"  Defect structures: {len(defect_data)}")

    sub_data = []
    substituents = ["Li", "Na", "K", "Ca", "Cu", "Ag", "Mg", "Zn", "Fe", "Ti",
                    "Al", "Cr", "Ni", "Co", "Mn", "V", "Pt", "Au", "Pd", "Ir",
                    "W", "Mo", "Nb", "Ta", "Zr", "Hf", "Ru", "Rh", "Os", "Re"]
    for entry in all_entries:
        try:
            first_sym = entry.formula.rstrip("0123456789")[:2]
            subs = engine.substitution_scan(entry.formula, first_sym, substituents)
            for s in subs:
                sub_data.append({
                    "base_formula": entry.formula,
                    "substituted_formula": s.formula,
                    "stability_score": round(s.stability_score, 4),
                    "stability_class": s.stability.value if hasattr(s.stability, "value") else str(s.stability),
                    "provenance": s.provenance,
                })
        except Exception:
            pass
    log(f"  Substitution variants: {len(sub_data)}")

    # Stoichiometric perturbations
    perturb_data = []
    for entry in all_entries:
        try:
            perts = engine.stoichiometric_perturbation(entry.formula, n_variants=10)
            for p in perts:
                perturb_data.append({
                    "base_formula": entry.formula,
                    "perturbed_formula": p.formula,
                    "stability_score": round(p.stability_score, 4),
                    "stability_class": p.stability.value if hasattr(p.stability, "value") else str(p.stability),
                    "provenance": p.provenance,
                })
        except Exception:
            pass
    log(f"  Stoichiometric perturbations: {len(perturb_data)}")

    # Write CSVs
    cat_csv = OUT_ROOT / "prong2_catalog.csv"
    with open(cat_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "formula", "crystal_class", "bravais", "space_group",
            "a_A", "density_gcc", "stability", "stability_score", "n_atoms"])
        w.writeheader()
        for e in all_entries:
            w.writerow({
                "formula": e.formula,
                "crystal_class": e.crystal_class.value if hasattr(e.crystal_class, "value") else str(e.crystal_class),
                "bravais": e.bravais.value if hasattr(e.bravais, "value") else str(e.bravais),
                "space_group": getattr(e, "space_group", ""),
                "a_A": round(getattr(e, "a", 0), 4),
                "density_gcc": round(getattr(e, "density", 0), 4),
                "stability": e.stability.value if hasattr(e.stability, "value") else str(e.stability),
                "stability_score": round(getattr(e, "stability_score", 0), 4),
                "n_atoms": e.n_atoms,
            })

    for name, data in [("prong2_supercells.csv", supercell_data),
                       ("prong2_defects.csv", defect_data),
                       ("prong2_substitutions.csv", sub_data),
                       ("prong2_perturbations.csv", perturb_data)]:
        if data:
            with open(OUT_ROOT / name, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=list(data[0].keys()))
                w.writeheader()
                w.writerows(data)

    elapsed = time.time() - t0
    summary = {
        "prong": "2_crystals",
        "catalog_entries": len(all_entries),
        "supercells": len(supercell_data),
        "defects": len(defect_data),
        "substitutions": len(sub_data),
        "perturbations": len(perturb_data),
        "elapsed_s": round(elapsed, 2),
    }
    log(f"PRONG 2 complete: {len(all_entries)} catalog + {len(supercell_data)} sc + "
        f"{len(defect_data)} def + {len(sub_data)} sub + {len(perturb_data)} pert in {elapsed:.1f}s")
    return {"summary": summary, "catalog": all_entries,
            "supercells": supercell_data, "defects": defect_data,
            "substitutions": sub_data, "perturbations": perturb_data}


# ===========================================================================
# PRONG 3 -- Metals & Macros
# ===========================================================================

def prong3_metals() -> Dict[str, Any]:
    log("=" * 70)
    log("PRONG 3: Metals & Macros -- starting")
    t0 = time.time()

    all_symbols = sorted(_METAL_DB.keys())
    log(f"  Metals in database: {len(all_symbols)}")

    metals_rows = []
    for sym in all_symbols:
        m = _METAL_DB[sym]
        metals_rows.append({
            "symbol": sym,
            "name": m.name,
            "formula": m.formula,
            "density_gcc": m.density,
            "E_GPa": m.elastic_modulus,
            "G_GPa": m.shear_modulus,
            "K_GPa": m.bulk_modulus,
            "nu": m.poisson_ratio,
            "Sy_MPa": m.yield_strength,
            "Su_MPa": m.ultimate_strength,
            "elong_pct": getattr(m, "elongation", 0),
            "HV": getattr(m, "vickers_hardness", 0),
            "k_WmK": m.thermal_conductivity,
            "sigma_e_MSm": getattr(m, "electrical_conductivity", 0),
            "CTE_umK": getattr(m, "cte", 0),
            "Cp_JkgK": getattr(m, "specific_heat", 0),
            "Tm_K": m.melting_point,
            "cost_USDkg": getattr(m, "cost_per_kg", 0),
            "corrosion_class": m.corrosion_class,
            "source": m.source,
        })

    all_objectives = list(MissionObjective)
    mission_results = {}
    for obj in all_objectives:
        spec = MissionSpec(objectives=[obj])
        ranked = run_mission(spec, top_n=len(all_symbols))
        mission_results[obj.value] = []
        for rc in ranked:
            mission_results[obj.value].append({
                "rank": rc.rank if hasattr(rc, "rank") else 0,
                "material": rc.material.formula,
                "name": rc.material.name,
                "score": round(rc.score, 4),
                "subscores": {k: round(v, 4) for k, v in rc.subscores.items()}
                             if hasattr(rc, "subscores") and rc.subscores else {},
            })
    log(f"  Mission objectives evaluated: {len(all_objectives)}")

    # Multi-objective missions
    multi_missions = [
        ("aerospace", [MissionObjective.LOW_MASS, MissionObjective.HIGH_STRENGTH, MissionObjective.HIGH_TEMPERATURE]),
        ("marine", [MissionObjective.CORROSION_RESISTANT, MissionObjective.HIGH_STRENGTH, MissionObjective.LOW_COST]),
        ("electrical", [MissionObjective.HIGH_CONDUCTIVITY_ELECTRICAL, MissionObjective.LOW_COST, MissionObjective.MACHINABLE]),
        ("tooling", [MissionObjective.HIGH_STIFFNESS, MissionObjective.HIGH_STRENGTH, MissionObjective.MACHINABLE]),
        ("thermal_mgmt", [MissionObjective.HIGH_CONDUCTIVITY_THERMAL, MissionObjective.LOW_EXPANSION, MissionObjective.LOW_MASS]),
    ]
    for label, objs in multi_missions:
        spec = MissionSpec(objectives=objs)
        ranked = run_mission(spec, top_n=len(all_symbols))
        mission_results[f"multi_{label}"] = []
        for rc in ranked:
            mission_results[f"multi_{label}"].append({
                "rank": rc.rank if hasattr(rc, "rank") else 0,
                "material": rc.material.formula,
                "name": rc.material.name,
                "score": round(rc.score, 4),
            })
    log(f"  Multi-objective missions: {len(multi_missions)}")

    alloy_compositions = [
        {"Cu": 0.7, "Zn": 0.3},    {"Cu": 0.6, "Zn": 0.4},
        {"Cu": 0.9, "Sn": 0.1},    {"Cu": 0.85, "Sn": 0.15},
        {"Cu": 0.88, "Al": 0.10, "Ni": 0.02},
        {"Fe": 0.98, "C": 0.02},   {"Fe": 0.96, "C": 0.04},
        {"Fe": 0.88, "Cr": 0.12},  {"Fe": 0.82, "Cr": 0.18},
        {"Fe": 0.7, "Cr": 0.2, "Ni": 0.1},
        {"Fe": 0.67, "Cr": 0.25, "Ni": 0.08},
        {"Fe": 0.74, "Cr": 0.18, "Ni": 0.08},
        {"Ti": 0.9, "Al": 0.06, "V": 0.04},
        {"Ti": 0.94, "Al": 0.06},
        {"Ti": 0.85, "V": 0.15},
        {"Ni": 0.7, "Cr": 0.15, "Fe": 0.15},
        {"Ni": 0.55, "Ti": 0.45},
        {"Ni": 0.76, "Cr": 0.16, "Fe": 0.08},
        {"Al": 0.95, "Cu": 0.03, "Mg": 0.02},
        {"Al": 0.90, "Zn": 0.06, "Mg": 0.02, "Cu": 0.02},
        {"Al": 0.93, "Si": 0.07},
        {"Co": 0.6, "Cr": 0.25, "W": 0.15},
        {"Pt": 0.9, "Ir": 0.1},
        {"Au": 0.75, "Cu": 0.125, "Ag": 0.125},
        {"W": 0.9, "Ni": 0.06, "Fe": 0.04},
        {"W": 0.97, "Re": 0.03},
        {"Zr": 0.975, "Sn": 0.015, "Fe": 0.01},
        {"Mg": 0.91, "Al": 0.09},
        {"Nb": 0.89, "Ti": 0.11},
        {"Mo": 0.5, "Re": 0.5},
        {"Cr": 0.8, "Mo": 0.2},
    ]
    alloy_rows = []
    for comp in alloy_compositions:
        try:
            alloy = AlloyEstimator.rule_of_mixtures(comp)
            if alloy:
                d = asdict(alloy)
                d["composition"] = json.dumps(comp)
                alloy_rows.append(d)
        except Exception:
            pass
    log(f"  Alloy estimates: {len(alloy_rows)}")

    heating_rows = []
    for sym in all_symbols:
        try:
            steps = simulate_heating(sym, 200, 4000, 100)
            for step in steps:
                heating_rows.append({
                    "material": sym,
                    "T_K": step.T,
                    "Cp_at_T": round(step.Cp_at_T, 4) if hasattr(step, "Cp_at_T") else 0,
                    "phase": step.phase if hasattr(step, "phase") else "",
                    "enthalpy_kJ": round(step.enthalpy_kJ, 4) if hasattr(step, "enthalpy_kJ") else 0,
                })
        except Exception:
            pass
    log(f"  Heating simulation points: {len(heating_rows)}")

    # Write CSVs
    for name, data in [("prong3_metals.csv", metals_rows),
                       ("prong3_alloys.csv", alloy_rows),
                       ("prong3_heating.csv", heating_rows)]:
        if data:
            with open(OUT_ROOT / name, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=list(data[0].keys()))
                w.writeheader()
                w.writerows(data)

    elapsed = time.time() - t0
    summary = {
        "prong": "3_metals",
        "metals": len(all_symbols),
        "missions_single": len(all_objectives),
        "missions_multi": len(multi_missions),
        "alloy_estimates": len(alloy_rows),
        "heating_points": len(heating_rows),
        "elapsed_s": round(elapsed, 2),
    }
    log(f"PRONG 3 complete: {len(metals_rows)} metals, {len(alloy_rows)} alloys, "
        f"{len(heating_rows)} heating pts in {elapsed:.1f}s")
    return {"summary": summary, "metals": metals_rows, "alloys": alloy_rows,
            "heating": heating_rows, "missions": mission_results}


# ===========================================================================
# PRONG 4 -- SmartSampling
# ===========================================================================

def prong4_sampling(seed: int = 42) -> Dict[str, Any]:
    log("=" * 70)
    log("PRONG 4: SmartSampling -- starting")
    t0 = time.time()

    all_samples = []

    steam_materials = ["H2O", "CO2", "NH3", "CH4", "R134a", "N2", "O2", "He", "Ar"]
    for formula in steam_materials:
        try:
            sampler = create_steam_sampler(formula, seed=seed)
            for tier in [SamplingTier.COARSE, SamplingTier.STANDARD,
                         SamplingTier.REFINED, SamplingTier.CRITICAL]:
                candidates = sampler.generate_candidates(n=300, tier=tier)
                ranked = sampler.rank_candidates(candidates)
                selected = sampler.select_top(ranked, n=150)
                for s in selected:
                    all_samples.append({
                        "domain": "steam",
                        "material": formula,
                        "tier": s.tier.value,
                        "total_score": round(s.total, 4),
                        "novelty": round(s.novelty, 4),
                        "objective": round(s.objective_match, 4),
                        "uncertainty": round(s.uncertainty, 4),
                        "plausibility": round(s.physical_plausibility, 4),
                        "instability": round(s.instability_interest, 4),
                        "cost": round(s.compute_affordability, 4),
                        "coords": json.dumps({k: round(v, 4) for k, v in s.coordinates.items()}),
                    })
            stress = sampler.stochastic_stress_test(n=100)
            for s in stress:
                all_samples.append({
                    "domain": "steam_stress",
                    "material": formula,
                    "tier": s.tier.value,
                    "total_score": round(s.total, 4),
                    "novelty": round(s.novelty, 4),
                    "objective": round(s.objective_match, 4),
                    "uncertainty": round(s.uncertainty, 4),
                    "plausibility": round(s.physical_plausibility, 4),
                    "instability": round(s.instability_interest, 4),
                    "cost": round(s.compute_affordability, 4),
                    "coords": json.dumps({k: round(v, 4) for k, v in s.coordinates.items()}),
                })
            log(f"  [sampling] {formula}: {len([s for s in all_samples if s['material']==formula])} total")
        except Exception as e:
            log(f"  [sampling] {formula} FAILED: {e}")

    try:
        cryst_sampler = create_crystal_sampler(seed=seed)
        for tier in [SamplingTier.COARSE, SamplingTier.STANDARD,
                     SamplingTier.REFINED, SamplingTier.CRITICAL]:
            candidates = cryst_sampler.generate_candidates(n=500, tier=tier)
            ranked = cryst_sampler.rank_candidates(candidates)
            selected = cryst_sampler.select_top(ranked, n=250)
            for s in selected:
                all_samples.append({
                    "domain": "crystal",
                    "material": "",
                    "tier": s.tier.value,
                    "total_score": round(s.total, 4),
                    "novelty": round(s.novelty, 4),
                    "objective": round(s.objective_match, 4),
                    "uncertainty": round(s.uncertainty, 4),
                    "plausibility": round(s.physical_plausibility, 4),
                    "instability": round(s.instability_interest, 4),
                    "cost": round(s.compute_affordability, 4),
                    "coords": json.dumps({k: round(v, 4) for k, v in s.coordinates.items()}),
                })
        log(f"  [sampling] crystal: {len([s for s in all_samples if s['domain']=='crystal'])} total")
    except Exception as e:
        log(f"  [sampling] crystal FAILED: {e}")

    alloy_element_sets = [
        ["Cu", "Zn", "Sn"], ["Fe", "Cr", "Ni"], ["Ti", "Al", "V"],
        ["Ni", "Co", "Cr", "W"], ["Al", "Cu", "Mg", "Zn"],
        ["Mo", "Re", "W"], ["Nb", "Ti", "Zr"], ["Pt", "Pd", "Rh"],
    ]
    for elems in alloy_element_sets:
        tag = "-".join(elems)
        try:
            alloy_sampler = create_alloy_sampler(elems, seed=seed)
            for tier in [SamplingTier.STANDARD, SamplingTier.REFINED, SamplingTier.CRITICAL]:
                candidates = alloy_sampler.generate_candidates(n=300, tier=tier)
                ranked = alloy_sampler.rank_candidates(candidates)
                selected = alloy_sampler.select_top(ranked, n=150)
                for s in selected:
                    all_samples.append({
                        "domain": f"alloy_{tag}",
                        "material": tag,
                        "tier": s.tier.value,
                        "total_score": round(s.total, 4),
                        "novelty": round(s.novelty, 4),
                        "objective": round(s.objective_match, 4),
                        "uncertainty": round(s.uncertainty, 4),
                        "plausibility": round(s.physical_plausibility, 4),
                        "instability": round(s.instability_interest, 4),
                        "cost": round(s.compute_affordability, 4),
                        "coords": json.dumps({k: round(v, 4) for k, v in s.coordinates.items()}),
                    })
            log(f"  [sampling] alloy {tag}: done")
        except Exception as e:
            log(f"  [sampling] alloy {tag} FAILED: {e}")

    samp_csv = OUT_ROOT / "prong4_sampling.csv"
    if all_samples:
        with open(samp_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(all_samples[0].keys()))
            w.writeheader()
            w.writerows(all_samples)

    elapsed = time.time() - t0
    summary = {
        "prong": "4_sampling",
        "total_samples": len(all_samples),
        "steam_samples": len([s for s in all_samples if s["domain"].startswith("steam")]),
        "crystal_samples": len([s for s in all_samples if s["domain"] == "crystal"]),
        "alloy_samples": len([s for s in all_samples if s["domain"].startswith("alloy")]),
        "elapsed_s": round(elapsed, 2),
    }
    log(f"PRONG 4 complete: {len(all_samples)} total samples in {elapsed:.1f}s")
    return {"summary": summary, "samples": all_samples}


# ===========================================================================
# PRONG 5 -- Golden Project
# ===========================================================================

def prong5_golden(seed: int = 42) -> Dict[str, Any]:
    log("=" * 70)
    log("PRONG 5: Golden Project -- starting")
    t0 = time.time()

    atlas = DiscoveryAtlas(output_dir=str(OUT_ROOT / "golden_project"))

    full_materials = ["Fe", "Al", "Ti", "Cu", "Ni", "W", "Co", "Cr", "Zr", "Nb",
                      "Mo", "Pt", "Au", "Ag", "Pd", "Mg", "Zn", "Sn", "Pb", "Mn",
                      "V", "Ta", "Hf", "Ru", "Rh", "Ir", "Os", "Re"]

    missions = [
        DiscoveryMission(
            mission_id="full_discovery_001",
            name="Full Discovery Atlas",
            mission_type=MissionType.FULL_DISCOVERY,
            description="Complete cross-pillar sweep",
            materials=full_materials,
            seed=seed,
            max_iterations=10,
            samples_per_iteration=100,
        ),
        DiscoveryMission(
            mission_id="thermo_deep_001",
            name="Deep Thermophysical Exploration",
            mission_type=MissionType.THERMO_EXPLORATION,
            description="Deep thermo sweep for all S1 materials",
            materials=["H2O", "CO2", "NH3", "CH4", "R134a", "N2", "O2", "He", "Ar"],
            seed=seed + 1,
            max_iterations=12,
            samples_per_iteration=120,
        ),
        DiscoveryMission(
            mission_id="crystal_search_001",
            name="Crystal Structure Search",
            mission_type=MissionType.CRYSTAL_SEARCH,
            description="Exhaustive crystal structure generation and ranking",
            seed=seed + 2,
            max_iterations=8,
            samples_per_iteration=80,
        ),
        DiscoveryMission(
            mission_id="alloy_opt_001",
            name="Alloy Optimization Campaign",
            mission_type=MissionType.ALLOY_OPTIMIZATION,
            description="Multi-objective alloy screening",
            targets={"objectives": ["high_strength", "high_temperature",
                                    "low_cost", "low_mass", "corrosion_resistant"]},
            constraints={"density": (1.0, 12.0), "melting_point": (500, 4000)},
            materials=full_materials[:15],
            seed=seed + 3,
            max_iterations=10,
            samples_per_iteration=100,
        ),
        DiscoveryMission(
            mission_id="process_eval_001",
            name="Heating Process Evaluation",
            mission_type=MissionType.PROCESS_EVALUATION,
            description="Evaluate heating profiles for all candidate materials",
            materials=full_materials,
            seed=seed + 4,
            max_iterations=8,
            samples_per_iteration=80,
        ),
        DiscoveryMission(
            mission_id="alloy_opt_002",
            name="Aerospace Alloy Campaign",
            mission_type=MissionType.ALLOY_OPTIMIZATION,
            description="Aerospace-focused alloy screening: low mass + high T",
            targets={"objectives": ["low_mass", "high_temperature", "high_strength"]},
            materials=["Ti", "Al", "Ni", "Co", "Cr", "Nb", "Mo", "W"],
            seed=seed + 5,
            max_iterations=8,
            samples_per_iteration=80,
        ),
    ]

    all_results = []
    for mission in missions:
        log(f"  Running mission: {mission.name} ({mission.mission_type.value})")
        try:
            results = atlas.run_mission(mission)
            all_results.extend(results)
            log(f"    -> {len(results)} results")
        except Exception as e:
            log(f"    -> FAILED: {e}")
            traceback.print_exc()

    atlas.export_all_csv()
    report_md = atlas.generate_master_report_md()
    report_path = OUT_ROOT / "golden_project" / "master_report.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_md)

    elapsed = time.time() - t0
    summary = {
        "prong": "5_golden",
        "missions": len(missions),
        "total_results": len(all_results),
        "elapsed_s": round(elapsed, 2),
    }
    log(f"PRONG 5 complete: {len(missions)} missions, {len(all_results)} results in {elapsed:.1f}s")
    return {"summary": summary, "results": all_results, "atlas": atlas}


# ===========================================================================
# XLSX CONSOLIDATION
# ===========================================================================

def build_xlsx(p1, p2, p3, p4, p5) -> str:
    log("=" * 70)
    log("XLSX CONSOLIDATION -- building multi-page workbook")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Sheet: Summary ---
    ws = wb.create_sheet("Summary")
    summaries = [p1["summary"], p2["summary"], p3["summary"],
                 p4["summary"], p5["summary"]]
    sum_headers = ["Prong", "Metric", "Value", "Elapsed (s)"]
    sum_rows = []
    for s in summaries:
        prong = s["prong"]
        elapsed = s["elapsed_s"]
        for k, v in s.items():
            if k not in ("prong", "elapsed_s"):
                sum_rows.append([prong, k, str(v), elapsed])
    _write_sheet(ws, sum_headers, sum_rows)

    # --- Sheet: Thermo Grid ---
    ws = wb.create_sheet("Thermo Grid")
    grid_headers = ["Material", "T (K)", "P (Pa)", "Phase", "h (J/kg)",
                    "s (J/kg K)", "v (m3/kg)", "Cp (J/kg K)", "Cv (J/kg K)",
                    "mu (Pa s)", "k (W/m K)", "Pr", "Quality x", "Z_comp"]
    grid_rows = []
    for formula in p1.get("formulas", []):
        for row in p1["grid"].get(formula, []):
            grid_rows.append([
                row["material"], row["T_K"], row["P_Pa"], row["phase"],
                row["h_Jkg"], row["s_JkgK"], row["v_m3kg"],
                row["cp_JkgK"], row["cv_JkgK"], row["mu_Pas"],
                row["k_WmK"], row["Pr"], row["quality_x"], row["Z_comp"],
            ])
    _write_sheet(ws, grid_headers, grid_rows)

    # --- Sheet: Saturation ---
    ws = wb.create_sheet("Saturation")
    sat_headers = ["Material", "T (K)", "P (Pa)", "Phase", "h (J/kg)",
                   "s (J/kg K)", "v (m3/kg)"]
    sat_rows = []
    for formula in p1.get("formulas", []):
        for row in p1["sat"].get(formula, []):
            sat_rows.append([
                row["material"], row["T_K"], row["P_Pa"], row["phase"],
                row["h_Jkg"], row["s_JkgK"], row["v_m3kg"],
            ])
    _write_sheet(ws, sat_headers, sat_rows)

    # --- Sheet: Crystal Catalog ---
    ws = wb.create_sheet("Crystal Catalog")
    cryst_headers = ["Formula", "Class", "Bravais", "Space Group", "a (A)",
                     "Density (g/cc)", "Stability", "Score", "N Atoms"]
    cryst_rows = []
    for e in p2.get("catalog", []):
        cryst_rows.append([
            e.formula,
            e.crystal_class.value if hasattr(e.crystal_class, "value") else str(e.crystal_class),
            e.bravais.value if hasattr(e.bravais, "value") else str(e.bravais),
            getattr(e, "space_group", ""),
            round(getattr(e, "a", 0), 4),
            round(getattr(e, "density", 0), 4),
            e.stability.value if hasattr(e.stability, "value") else str(e.stability),
            round(getattr(e, "stability_score", 0), 4),
            e.n_atoms,
        ])
    _write_sheet(ws, cryst_headers, cryst_rows)

    # --- Sheet: Supercells ---
    ws = wb.create_sheet("Supercells")
    sc_headers = ["Base Formula", "Supercell", "N Atoms", "Volume (A3)",
                  "Lattice", "Class", "Provenance"]
    sc_rows = [[d["base_formula"], d["supercell"], d["n_atoms"],
                d["volume_A3"], d["lattice"], d["crystal_class"],
                d["provenance"]] for d in p2.get("supercells", [])]
    _write_sheet(ws, sc_headers, sc_rows)

    # --- Sheet: Defects ---
    ws = wb.create_sheet("Defects")
    def_headers = ["Base Formula", "Defect Type", "N Atoms", "Volume (A3)", "Provenance"]
    def_rows = [[d["base_formula"], d["defect_type"], d["n_atoms"],
                 d["volume_A3"], d["provenance"]] for d in p2.get("defects", [])]
    _write_sheet(ws, def_headers, def_rows)

    # --- Sheet: Substitutions ---
    ws = wb.create_sheet("Substitutions")
    sub_headers = ["Base Formula", "Substituted", "Stability Score",
                   "Stability Class", "Provenance"]
    sub_rows = [[d["base_formula"], d["substituted_formula"],
                 d["stability_score"], d["stability_class"],
                 d["provenance"]] for d in p2.get("substitutions", [])]
    _write_sheet(ws, sub_headers, sub_rows)

    # --- Sheet: Perturbations ---
    ws = wb.create_sheet("Perturbations")
    pert_headers = ["Base Formula", "Perturbed", "Stability Score",
                    "Stability Class", "Provenance"]
    pert_rows = [[d["base_formula"], d["perturbed_formula"],
                  d["stability_score"], d["stability_class"],
                  d["provenance"]] for d in p2.get("perturbations", [])]
    _write_sheet(ws, pert_headers, pert_rows)

    # --- Sheet: Metals ---
    ws = wb.create_sheet("Metals")
    if p3.get("metals"):
        met_headers = list(p3["metals"][0].keys())
        met_rows = [list(m.values()) for m in p3["metals"]]
        _write_sheet(ws, met_headers, met_rows)

    # --- Sheet: Alloys ---
    ws = wb.create_sheet("Alloys")
    if p3.get("alloys"):
        alloy_headers = list(p3["alloys"][0].keys())
        alloy_rows = [list(a.values()) for a in p3["alloys"]]
        _write_sheet(ws, alloy_headers, alloy_rows)

    # --- Sheet: Heating ---
    ws_heat = wb.create_sheet("Heating")
    if p3.get("heating"):
        heat_headers = list(p3["heating"][0].keys())
        heat_rows = [list(h.values()) for h in p3["heating"]]
        _write_sheet(ws_heat, heat_headers, heat_rows)
        try:
            _add_line_chart(ws_heat, "Cp vs Temperature (all metals)", x_col=2, y_cols=[3])
        except Exception:
            pass

    # --- Sheet per mission objective ---
    for obj_name, rankings in p3.get("missions", {}).items():
        sheet_name = f"M_{obj_name[:26]}"
        ws = wb.create_sheet(sheet_name)
        if rankings:
            m_headers = ["Rank", "Material", "Name", "Score"]
            m_rows = [[r.get("rank", 0), r["material"], r["name"], r["score"]] for r in rankings]
            _write_sheet(ws, m_headers, m_rows)

    # --- Sheet: Sampling ---
    ws = wb.create_sheet("Sampling")
    if p4.get("samples"):
        samp_headers = list(p4["samples"][0].keys())
        samp_rows = [list(s.values()) for s in p4["samples"]]
        _write_sheet(ws, samp_headers, samp_rows)

    # --- Sheet: Golden Results ---
    ws = wb.create_sheet("Golden Results")
    gold_headers = ["Mission ID", "Pillar", "Material", "Score", "Rank",
                    "Provenance", "Properties"]
    gold_rows = []
    for r in p5.get("results", []):
        props_str = json.dumps(r.properties, default=str)[:200] if hasattr(r, "properties") else ""
        gold_rows.append([
            r.mission_id, r.pillar, r.material, round(r.score, 4),
            r.rank, r.provenance, props_str,
        ])
    _write_sheet(ws, gold_headers, gold_rows)

    # Save
    xlsx_path = str(OUT_ROOT / f"VSEPR_SIM_5prong_{TIMESTAMP}.xlsx")
    wb.save(xlsx_path)
    log(f"XLSX saved: {xlsx_path}")
    log(f"  Sheets: {len(wb.sheetnames)}")
    log(f"  Sheet names: {wb.sheetnames}")
    return xlsx_path


# ===========================================================================
# FINAL REPORT + EMAIL
# ===========================================================================

def write_final_report(summaries, xlsx_path) -> str:
    report_path = str(OUT_ROOT / "final_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("VSEPR-SIM 5-Prong Large Data Creation -- Final Report\n")
        f.write("=" * 60 + "\n")
        f.write(f"Timestamp: {TIMESTAMP}\n\n")
        total_time = sum(s.get("elapsed_s", 0) for s in summaries)
        f.write(f"Total simulation time: {total_time:.1f}s\n\n")
        for s in summaries:
            f.write(f"--- {s['prong']} ---\n")
            for k, v in s.items():
                f.write(f"  {k}: {v}\n")
            f.write("\n")
        f.write(f"\nXLSX output: {xlsx_path}\n")
        f.write(f"Output directory: {OUT_ROOT}\n")
    return report_path


def write_summary_json(summaries, xlsx_path) -> str:
    json_path = str(OUT_ROOT / "summary.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": TIMESTAMP,
            "prongs": summaries,
            "xlsx_path": xlsx_path,
            "output_dir": str(OUT_ROOT),
        }, f, indent=2)
    return json_path


def email_results(
    smtp_server: str, smtp_port: int, sender_email: str,
    app_password: str, recipient_email: str, subject: str,
    body: str, attachments: list | None = None,
) -> bool:
    import smtplib
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["From"] = sender_email
    msg["To"] = recipient_email
    msg["Subject"] = subject
    msg.set_content(body)
    for file_path in attachments or []:
        path = Path(file_path)
        if not path.exists():
            continue
        data = path.read_bytes()
        msg.add_attachment(data, maintype="application", subtype="octet-stream", filename=path.name)
    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as smtp:
            smtp.login(sender_email, app_password)
            smtp.send_message(msg)
        log("[email] Sent successfully.")
        return True
    except Exception as e:
        log(f"[email] FAILED: {e}")
        return False


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    log("=" * 70)
    log("VSEPR-SIM 5-Prong Large Data Creation Engine")
    log(f"Output: {OUT_ROOT}")
    log(f"Timestamp: {TIMESTAMP}")
    log("Strategy: ACCURACY over speed | parallelism for throughput")
    log("=" * 70)

    t_total = time.time()

    p1 = prong1_thermo(n_T=80, n_P=80, sat_pts=200)
    p2 = prong2_crystals(seed=42)
    p3 = prong3_metals()
    p4 = prong4_sampling(seed=42)
    p5 = prong5_golden(seed=42)

    xlsx_path = build_xlsx(p1, p2, p3, p4, p5)

    summaries = [p1["summary"], p2["summary"], p3["summary"],
                 p4["summary"], p5["summary"]]
    report_path = write_final_report(summaries, xlsx_path)
    json_path = write_summary_json(summaries, xlsx_path)

    total_elapsed = time.time() - t_total
    log("=" * 70)
    log(f"ALL 5 PRONGS COMPLETE in {total_elapsed:.1f}s")
    log(f"  XLSX: {xlsx_path}")
    log(f"  Report: {report_path}")
    log(f"  JSON: {json_path}")
    log(f"  Log: {_LOG_FILE}")
    log("=" * 70)

    # Email -- configure sender credentials to enable
    email_cfg = {
        "smtp_server": "smtp.gmail.com",
        "smtp_port": 465,
        "sender_email": "CONFIGURE_SENDER@gmail.com",
        "app_password": "CONFIGURE_APP_PASSWORD",
        "recipient_email": "liammichaelcrosby@outlook.com",
        "subject": f"VSEPR-SIM 5-Prong Run Complete -- {TIMESTAMP}",
        "body": (
            f"VSEPR-SIM 5-Prong Large Data Creation run finished.\n"
            f"Total time: {total_elapsed:.1f}s\n\n"
            + "\n".join(f"  {s['prong']}: {s}" for s in summaries)
            + f"\n\nAttached: XLSX workbook, final report, summary JSON."
        ),
        "attachments": [xlsx_path, report_path, json_path],
    }

    if email_cfg["sender_email"] != "CONFIGURE_SENDER@gmail.com":
        email_results(**email_cfg)
    else:
        log("[email] Skipped -- set sender_email + app_password to enable delivery")
        log(f"[email] Recipient configured: {email_cfg['recipient_email']}")

    return {"xlsx": xlsx_path, "report": report_path, "json": json_path,
            "log": str(_LOG_FILE), "summaries": summaries,
            "elapsed_s": round(total_elapsed, 2)}


if __name__ == "__main__":
    main()

